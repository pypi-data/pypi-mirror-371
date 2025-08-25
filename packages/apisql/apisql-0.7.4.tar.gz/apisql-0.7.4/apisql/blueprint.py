import codecs
import os
import csv
import urllib
import tempfile
from io import StringIO

from flask import Blueprint, Response, request, send_file, abort
from flask_jsonpify import jsonpify

import openpyxl

from .controllers import Controllers
from .logger import logger, logging

MAX_ROWS = int(os.environ.get('APISQL__MAX_ROWS', 100))
CONNECTION_STRING = os.environ.get('APISQL__DATABASE_URL')


class APISQLBlueprint(Blueprint):

    def __init__(self, connection_string=CONNECTION_STRING, engine=None, max_rows=MAX_ROWS, debug=False, cache=None, external_url=None):
        super().__init__('apisql', 'apisql')
        self.controllers = Controllers(
            connection_string, max_rows, debug, engine
        )
        if debug:
            logger.setLevel(logging.DEBUG)
        else:
            logger.setLevel(logging.INFO)
        self.max_rows = max_rows

        self.add_url_rule(
            '/query',
            'query',
            self.query,
            methods=['GET', 'POST']
        )
        self.add_url_rule(
            '/download',
            'download',
            self.download,
            methods=['GET', 'POST']
        )
        self.cache = cache
        self.external_url = external_url

    def query(self):
        status = []
        results = dict(total=0, rows=[])
        if not self.detect_bot():
            try:
                num_rows = int(request.values.get('num_rows', self.max_rows))
                page_size = int(request.values.get('page_size', num_rows))
                page = int(request.values.get('page', 0))
            except Exception:
                abort(400)
            sql = request.values.get('query')
            try:
                sql = codecs.decode(sql.encode('ascii'), 'base64').decode('utf8')
                status.append('base64 query')
            except Exception:
                status.append('plaintext query')
                pass
            key = '::'.join([sql, str(num_rows), str(page_size), str(page)])
            if self.cache is not None:
                results = self.cache.get(key)
            else:
                status.append('no cache')
            if results is None or self.cache is None:
                if self.cache is not None:
                    status.append('cache miss')
                results = self.controllers.query_db(sql, num_rows=num_rows, page_size=page_size, page=page)
                if 'download_url' in results and self.external_url:
                    results['download_url'] = self.external_url + results['download_url']
                if self.cache is not None:
                    self.cache.set(key, results)
            else:
                status.append('cache hit')
        else:
            status.append('Bot detected')
        return jsonpify(dict(**results, status=status))

    def download(self):
        format = request.values.get('format', 'xlsx')

        file_name = request.values.get('filename')
        # Create a default value here in case this parameter is not provided
        if file_name is None:
            file_name = 'query-results'

        formatters = request.values.get('headers').split(';')

        if format not in ('csv', 'xlsx'):
            abort(400)

        if self.detect_bot():
            headers = {
                'Content-Type': 'text/csv',
                'Content-Disposition': 'attachment; filename=bot-detected.csv'
            }
            return Response('', content_type='text/csv', headers=headers)

        mime = {
            'csv': 'text/csv',
            'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }[format]

        sql = request.values.get('query')
        try:
            try:
                sql = codecs.decode(sql.encode('ascii'), 'base64').decode('utf8')
            except Exception:
                pass
            results = self.controllers.query_db_streaming(sql, formatters)

            if format == 'csv':
                def generate():
                    buffer = StringIO()
                    writer = csv.writer(buffer)
                    for row in results:
                        writer.writerow(row)
                        pos = buffer.tell()
                        buffer.seek(0)
                        ret = buffer.read(pos)
                        buffer.seek(0)
                        yield ret

                # Encode the filename in utf-8 and url encoding
                file_name_utf8_encoded = file_name.encode('utf-8')
                file_name_url_encoded = urllib.parse.quote(file_name_utf8_encoded)

                headers = {
                    'Content-Type': mime,
                    'Content-Disposition': 'attachment; filename=' + file_name_url_encoded + '.csv'
                }
                return Response(generate(),
                                content_type='text/csv', headers=headers)
            if format == 'xlsx':
                with tempfile.NamedTemporaryFile(mode='w+b', suffix='.xlsx') as out:
                    try:
                        workbook = openpyxl.Workbook()
                        worksheet = workbook.active
                        for i, row in enumerate(results):
                            for j, v in enumerate(row):
                                if v is not None:
                                    try:
                                        worksheet.cell(row=i+1, column=j+1, value=float(v))
                                    except ValueError:
                                        worksheet.cell(row=i+1, column=j+1, value=str(v))
                        workbook.save(out.name)
                    finally:
                        workbook.close()
                    return send_file(out.name, mimetype=mime, as_attachment=True, download_name=file_name + '.xlsx')
        except Exception as e:
            logger.error('Error downloading query: %s', str(e))
            abort(400, description=f'Error downloading query: {repr(e)} \nsql: {sql}')

    def detect_bot(self):
        if request.user_agent.browser in ('google', 'aol', 'baidu', 'bing', 'yahoo'):
            logger.info('Bot detected %s: %s', request.user_agent.string, request.user_agent.browser)
        elif any(x in request.user_agent.string.lower() for x in ('applebot', 'yandexbot', 'petalbot')):
            logger.info('Bot detected %s: %s', request.user_agent.string, request.user_agent.browser)
        else:
            return False
        return True
