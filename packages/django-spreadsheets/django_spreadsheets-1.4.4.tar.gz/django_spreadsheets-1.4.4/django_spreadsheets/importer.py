# Standard Library
import hashlib
import operator
from collections import defaultdict
from datetime import date, datetime

# Django
from django.core.exceptions import ValidationError
from django.core.validators import URLValidator
from django.utils.translation import gettext as _

# Third party
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

# Local application / specific library imports
from .constants import VERSION_CUSTOM_PROPERTY_NAME, YES_NO_CHOICES
from .errors import (
    ConfigVersionWarning,
    ImporterColumnHeaderError,
    ImporterRequiredDataError,
    ImporterRequiredIfDataError,
    ImporterRequiredUnlessDataError,
    ImporterSheetIgnored,
    ImporterSheetMissing,
    ImporterTypeError,
    ImporterUniquenessError,
)
from .shortcuts import extract_rows, xlsx_to_bool


class Importer:
    def __init__(self, config, file=None, matched_data=None, **kwargs):
        self.config = config
        self.load_errors = []
        self._import_errors = []
        self._import_warnings = []
        self.match_errors = {}
        self.matched_data = matched_data
        # A boolean to indicate wether the importer has run validation in order
        # to differentiate the absence of errors/warnings after validation from
        # the fact that the validation has not been run yet
        self.has_run_validation = False
        self.is_data_valid = False
        self.file = file

    def set_file(self, file):
        self.file = file

    # Workbook utils
    def load_file(
        self,
        data_only,
    ):
        if self.file is None:
            self.load_errors.append(ValidationError(_("No file were provided")))

        try:
            self.workbook = load_workbook(filename=self.file, data_only=data_only)
        except InvalidFileException as e:
            self.load_errors.append(e)

    def set_matched_data(self, data):
        self.matched_data = data

    def clean_string(self, string):
        # Replace "\n" by a space to avoid concatenating two words. For example avoid:
        # "this is a
        #  sentence" to become "this is asentence"
        if string is None:
            return ""

        string = str(string)
        return string.replace("\n", " ").strip()

    def get_cell_value_key(self, sheet, column_key, cell_value):
        return hashlib.md5(
            f"title_{sheet.title}_column_{column_key}_value_{cell_value}".encode()
        ).hexdigest()

    def get_valid_sheets(self):
        """
        Return worksheet instances that are valid for the current user
        ie: ignore the config sheet and sheets that are not described in sheets_config
        """
        return filter(
            lambda sheet: sheet.title in self.config.get_config_sheets_name()
            and sheet.sheet_state != sheet.SHEETSTATE_HIDDEN,
            self.workbook.worksheets,
        )

    def get_visible_sheets(self):
        """Return worksheet instances that are not hidden"""

        return filter(
            lambda sheet: sheet.sheet_state != sheet.SHEETSTATE_HIDDEN,
            self.workbook.worksheets,
        )

    # Workbook content validation
    def validate_data(self):
        # Load file, with data_only False to keep formulae
        self.load_file(data_only=True)

        # Continue only if there are no errors when loading the file
        if self.workbook is not None and not self.load_errors:
            version_warning = self.validate_config_version()
            if version_warning:
                self._import_warnings.extend(version_warning)

            sheets_errors = self.validate_sheets()
            if sheets_errors:
                self._import_errors.extend(sheets_errors)

            columns_headers_errors = self.validate_columns_headers()
            if columns_headers_errors:
                self._import_errors.extend(columns_headers_errors)

            required_data_errors = self.validate_required_data()
            if required_data_errors:
                self._import_errors.extend(required_data_errors)

            uniqueness_errors = self.validate_uniqueness()
            if uniqueness_errors:
                self._import_errors.extend(uniqueness_errors)

            types_errors = self.validate_types()
            if types_errors:
                self._import_errors.extend(types_errors)

            match_errors = self.validate_choices()
            if match_errors:
                self.match_errors.update(match_errors)

            self.workbook.close()

        self.is_data_valid = (
            not self.load_errors
            and not self._import_errors
            and not self._import_warnings
            and not match_errors
        )
        self.has_run_validation = True

    # Validators retrieving some errors
    def validate_config_version(self):
        try:
            version_found = self.workbook.custom_doc_props[
                VERSION_CUSTOM_PROPERTY_NAME
            ].value
        except KeyError:
            version_found = None

        if version_found != self.config.version:
            version_warning = ConfigVersionWarning(
                expected=self.config.version, found=version_found
            )
            return [version_warning]
        return []

    def validate_sheets(self):
        # Validate that all sheets in config exist in imported file
        missing_sheets = [
            ImporterSheetMissing(sheet_name)
            for sheet_name in self.config.get_config_sheets_name()
            if sheet_name not in [sheet.title for sheet in self.get_valid_sheets()]
        ]

        # Validate that all sheets in imported file exist in config
        ignored_sheets = [
            ImporterSheetIgnored(sheet.title, self.config.get_config_sheets_name())
            for sheet in self.get_visible_sheets()
            if sheet.title not in self.config.get_config_sheets_name()
        ]

        return missing_sheets + ignored_sheets

    def validate_columns_headers(self):
        columns_headers_errors = []
        for sheet in self.get_valid_sheets():
            columns_configs = self.config.get_sheet_columns_configs_list(sheet)
            for row in sheet.iter_rows(max_row=1):
                for cell in row:
                    try:
                        column_key, column_config = columns_configs[cell.col_idx - 1]
                        column_header = column_config.get("name", column_key)
                        column_header = self.config.get_column_header_with_suffix(
                            column_header, column_config
                        )
                    except IndexError:
                        break

                    if cell.value != column_header:
                        columns_headers_errors.append(
                            ImporterColumnHeaderError(
                                column_header=cell.value,
                                expected_column_header=column_header,
                                column=cell.column,
                                row=cell.row,
                                sheet_name=sheet.title,
                            )
                        )

        return columns_headers_errors

    def validate_required_data(self):
        required_data_errors = []

        for sheet in self.get_valid_sheets():
            columns_configs = self.config.get_sheet_columns_configs_list(sheet)
            must_break_row_loop = False

            for row in sheet.iter_rows(
                min_row=2,
            ):
                # Skip all rows which have no value or only formulae
                if all(cell.data_type == "f" or cell.value is None for cell in row):
                    continue

                # Check required cells that are empty and
                # required_unless cells that are empty together
                for cell in row:
                    is_cell_empty = cell.data_type != "f" and cell.value is None

                    try:
                        column_config = columns_configs[cell.column - 1][1]
                        column_key = columns_configs[cell.column - 1][0]
                        column_header = column_config.get("name", column_key)
                    except IndexError:
                        must_break_row_loop = True
                        break

                    required = column_config.get("required", False)

                    # 'Required unless' handling
                    required_unless_column_key = column_config.get("required_unless")
                    required_unless_column_index = None
                    if required_unless_column_key is not None:
                        try:
                            required_unless_column_index = list(
                                self.config.sheets_config[sheet.title].keys()
                            ).index(required_unless_column_key)
                        except ValueError:
                            raise ValueError(
                                f"The `required_unless` configuration field of column `{column_header}` references an unknown column: `{required_unless_column_key}`"
                            )
                        else:
                            required_unless_column_header = columns_configs[
                                required_unless_column_index
                            ][1].get("name", required_unless_column_key)
                            required_unless_column_letter = row[
                                required_unless_column_index
                            ].column_letter

                    # 'Required if' handling
                    required_if_column_key = column_config.get("required_if")
                    required_if_column_index = None
                    if required_if_column_key is not None:
                        try:
                            required_if_column_index = list(
                                self.config.sheets_config[sheet.title].keys()
                            ).index(required_if_column_key)
                        except ValueError:
                            raise ValueError(
                                f"The `required_if` configuration field of column `{column_header}` references an unknown column: `{required_if_column_key}`"
                            )
                        else:
                            required_if_column_header = columns_configs[
                                required_if_column_index
                            ][1].get("name", required_if_column_key)
                            required_if_column_letter = row[
                                required_if_column_index
                            ].column_letter

                    if is_cell_empty:
                        if required:
                            required_data_errors.append(
                                ImporterRequiredDataError(
                                    column_header=column_header,
                                    column=cell.column,
                                    row=cell.row,
                                    sheet_name=sheet.title,
                                )
                            )
                        elif required_unless_column_key:
                            is_unless_cell_empty = (
                                row[required_unless_column_index].value is None
                                and row[required_unless_column_index].data_type != "f"
                            )
                            if is_unless_cell_empty:
                                required_data_errors.append(
                                    ImporterRequiredUnlessDataError(
                                        required_unless_column_header=required_unless_column_header,
                                        required_unless_column_letter=required_unless_column_letter,
                                        column_header=column_header,
                                        column=cell.column,
                                        row=cell.row,
                                        sheet_name=sheet.title,
                                    )
                                )
                        elif required_if_column_key:
                            is_if_cell_empty = (
                                row[required_if_column_index].value is None
                                and row[required_if_column_index].data_type != "f"
                            )
                            if not is_if_cell_empty and is_cell_empty:
                                required_data_errors.append(
                                    ImporterRequiredIfDataError(
                                        required_if_column_header=required_if_column_header,
                                        required_if_column_letter=required_if_column_letter,
                                        column_header=column_header,
                                        column=cell.column,
                                        row=cell.row,
                                        sheet_name=sheet.title,
                                    )
                                )

                if must_break_row_loop:
                    break

        return required_data_errors

    def validate_uniqueness(self):
        """Validate that multiple cells in a row are unique together in the workbook."""
        # TODO: not implemented
        # TODO: this method depend on cols indexes instead of cols names

        uniqueness_errors = []

        def check_uniqueness(sheet, cols_indexes=()):
            """Validate that multiple cells in a row are unique together in the sheet

            Args:
                sheet: the sheet to validate
                cols_indexes (tuple, optional): a list of 1 based index of columns that must be unique together. Defaults to [].
            """
            unique_values = {}
            min_row = 2  # Skip header
            max_col = max(cols_indexes)

            line_number = min_row

            for row in sheet.iter_rows(min_row=2, max_col=max_col, values_only=True):
                values = tuple(
                    value for index, value in enumerate(row, 1) if index in cols_indexes
                )
                key = hash(values)
                if all(values) and key in unique_values:
                    uniqueness_errors.append(
                        ImporterUniquenessError(
                            line_number=line_number,
                            is_duplicate_of=unique_values[key],
                            sheet_name=sheet.title,
                        )
                    )
                else:
                    unique_values[key] = line_number
                line_number += 1

        for sheet in self.get_valid_sheets():
            # check_uniqueness(sheet, (1, 3))  # TODO: handle unique_together = ("A_COLUMN_IDENTIFIER", "ANOTHER_COLUMN_IDENTIFIER") in SheetConfig Meta class
            ...

        return uniqueness_errors

    def validate_types(self):
        """
        Validate that each column contains data of correct type (date, positive integer, ...)
        """

        types_errors = []

        for sheet in self.get_valid_sheets():
            columns_configs = self.config.get_sheet_columns_configs_list(sheet)
            must_break_row_loop = False

            for row in sheet.iter_rows(
                min_row=2,
            ):
                # Skip all rows which have no value or only formulae
                if all(cell.data_type == "f" or cell.value is None for cell in row):
                    continue

                for cell in row:
                    is_cell_empty = cell.data_type != "f" and cell.value is None
                    if not is_cell_empty:
                        try:
                            column_config = columns_configs[cell.column - 1][1]
                            column_key = columns_configs[cell.column - 1][0]
                            column_header = column_config.get("name", column_key)
                        except IndexError:
                            must_break_row_loop = True
                            break

                        validation = column_config.get("validation")
                        if validation is not None:
                            validation_type = validation["type"]

                            if validation_type == "whole":
                                if not isinstance(cell.value, int):
                                    types_errors.append(
                                        ImporterTypeError(
                                            expected_type="nombre",
                                            column=cell.column,
                                            row=cell.row,
                                            sheet_name=sheet.title,
                                            column_header=column_header,
                                            error_message=validation["error"],
                                        )
                                    )

                            elif validation_type == "url":
                                try:
                                    URLValidator()(self.clean_string(cell.value))
                                except ValidationError:
                                    types_errors.append(
                                        ImporterTypeError(
                                            expected_type="URL",
                                            column=cell.column,
                                            row=cell.row,
                                            sheet_name=sheet.title,
                                            column_header=column_header,
                                            error_message=validation["error"],
                                        )
                                    )

                            elif validation_type == "date":
                                if not isinstance(cell.value, (date, datetime)):
                                    types_errors.append(
                                        ImporterTypeError(
                                            expected_type="Date",
                                            column=cell.column,
                                            row=cell.row,
                                            sheet_name=sheet.title,
                                            column_header=column_header,
                                            error_message=validation["error"],
                                        )
                                    )

                if must_break_row_loop:
                    break

        return types_errors

    def validate_choices(self):
        """
        For each value of each column with a "list" validation, check
        that we find an object in the database matching the value in the
        spreadsheet.
        If we can't find a single `iexact` match, we store a match error for
        this value.
        """
        match_errors = {}

        for sheet in self.get_valid_sheets():
            columns_configs = self.config.get_sheet_columns_configs_list(sheet)
            must_break_row_loop = False

            for row in sheet.iter_rows(
                min_row=2,
            ):
                # Skip all rows which have no value or only formulae
                if all(cell.data_type == "f" or cell.value is None for cell in row):
                    continue

                else:
                    for cell in row:
                        try:
                            column_config = columns_configs[cell.column - 1][1]
                            column_key = columns_configs[cell.column - 1][0]
                            column_header = column_config.get("name", column_key)
                        except IndexError:
                            # No validation config found for this column, ignore it
                            continue

                        if (
                            "validation" in column_config
                            and column_config["validation"]["type"] == "list"
                        ):
                            validation_config = column_config["validation"]
                            is_required = column_config.get("required", False)

                            # Don't raise an error if field is not required and not filled
                            if not is_required and cell.value in (None, ""):
                                continue

                            choices = validation_config.get("formula1")
                            if callable(choices):
                                choices = choices()

                            try:
                                next(
                                    choice
                                    for choice in choices
                                    if choice.lower()
                                    == self.clean_string(cell.value).lower()
                                )

                            except StopIteration:
                                key = self.get_cell_value_key(
                                    sheet, column_key, cell.value
                                )

                                # Ignore if a match error already exists for this column
                                if key not in self.match_errors:
                                    match_errors[key] = {
                                        "label": _(
                                            'Sheet "{sheet_title}", column "{column_header}" ({column_letter}) select an existing value corresponding to "{cell_value}"'
                                        ).format(
                                            sheet_title=sheet.title,
                                            column_header=column_header,
                                            column_letter=cell.column_letter,
                                            cell_value=cell.value,
                                        ),
                                        "choices": [
                                            (
                                                "",
                                                _("Please select an existing element."),
                                            )
                                        ]
                                        + [(choice, choice) for choice in choices],
                                    }

                if must_break_row_loop:
                    break
        return match_errors

    # Workbook content extraction
    @property
    def values(self):
        """Return file values in a dict per sheet, replacing original invalid values with match_form cleaned values

        Args:
            match_form MatchForm: A valid MatchForm instance filled by the user

        Return:
            A dictionary with a key for each sheet, containing a list of dictionaries (a dictionary for each row).
        """

        # Re-load file, with data_only True to convert formulae to values
        self.load_file(data_only=True)

        values = defaultdict(list)

        for sheet in self.get_valid_sheets():
            columns_configs = self.config.get_sheet_columns_configs(sheet)
            columns_keys = list(columns_configs.keys())

            for row_number, row_values in extract_rows(sheet, columns_keys, min_row=2):
                for column_key, value in row_values.items():
                    validation_config = columns_configs[column_key].get("validation")
                    cell_key = self.get_cell_value_key(sheet, column_key, value)

                    # Replace unknown value in sheet by the matched value chosen by the user
                    if self.matched_data and cell_key in self.matched_data:
                        value = self.matched_data[cell_key]

                    # If the cell is a boolean choice
                    if (
                        validation_config is not None
                        and validation_config["type"] == "list"
                        and validation_config["formula1"] == YES_NO_CHOICES
                        and value in YES_NO_CHOICES
                    ):
                        value = xlsx_to_bool(value)

                    row_values[column_key] = value

                values[sheet.title].append(row_values)

        self.workbook.close()

        return values

    @property
    def import_errors(self):
        self._import_errors.sort(key=operator.attrgetter("title"))
        self._import_errors.sort(key=operator.attrgetter("order"))
        self._import_errors.sort(key=operator.attrgetter("sheet_name"))
        return self._import_errors

    @property
    def import_warnings(self):
        self._import_warnings.sort(key=operator.attrgetter("title"))
        self._import_warnings.sort(key=operator.attrgetter("order"))
        return self._import_warnings
