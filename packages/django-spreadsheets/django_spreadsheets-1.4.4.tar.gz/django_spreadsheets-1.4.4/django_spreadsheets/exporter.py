# Standard Library
from tempfile import NamedTemporaryFile

# Django
from django.utils.timezone import localtime, now

# Third party
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.packaging.custom import IntProperty
from openpyxl.styles import Alignment

# Local application / specific library imports
from .conf import settings as local_settings
from .constants import DEFAULT_HEADER_STYLE_NAME, VERSION_CUSTOM_PROPERTY_NAME
from .validation import ValidationMixin


class Exporter(ValidationMixin):
    def __init__(self, config, no_data=False, **kwargs) -> None:
        self.config = config
        self.no_data = no_data
        self.workbook = Workbook()
        self.set_version()
        self.init_named_styles()

    def process(self):
        # Do the export things here
        for index, (sheet_name, sheet_config) in enumerate(
            self.config.sheets_config.items()
        ):
            sheet = self.workbook.create_sheet(title=sheet_name, index=index)

            columns_count = self.create_header_row(sheet)

            # Data validation
            validations = self.setup_validations(
                sheet,
            )

            # Fill data in rows
            if not self.no_data:
                self.write_rows_data_to_sheet(sheet, sheet_config.get_data_to_export())

            # Apply validations
            for column_cursor in range(1, columns_count + 1):
                self.apply_validation_on_column(validations, column_cursor, sheet)

            self.set_columns_width(sheet, max_width=52)
            self.set_rows_height(sheet)
            self.set_column_styles(sheet, columns_count)

    def get_file_name(self):
        date = localtime(now()).strftime("%Y-%m-%d %Hh%Mm%Ss")
        return local_settings.EXPORT_FILE_NAME.format(
            date=date, config_name=str(self.config)
        )

    def save_to_file(self):
        self.process()

        # Remove the default sheet named "Sheet"
        self.workbook.remove(self.workbook["Sheet"])

        return self.workbook.save(self.get_file_name())

    def save_to_stream(self):
        self.process()

        # Remove the default sheet named "Sheet"
        self.workbook.remove(self.workbook["Sheet"])

        with NamedTemporaryFile(delete=False) as tmp:
            self.workbook.save(tmp.name)
            tmp.seek(0)
            return tmp.read()

    def create_header_row(self, sheet):
        columns_configs = self.config.get_sheet_columns_configs(sheet)

        column_cursor = 1
        for column_key, column_config in columns_configs.items():
            cell = sheet.cell(
                row=1,
                column=column_cursor,
                value=self.config.get_column_header_with_suffix(
                    column_config.get("name", column_key), column_config
                ),
            )
            cell.style = column_config.get("header_style", DEFAULT_HEADER_STYLE_NAME)

            comment = column_config.get("comment")
            if comment is not None:
                cell.comment = Comment(text=comment, author=self.config.comment_author)

            column_cursor += 1

        sheet.freeze_panes = "A2"

        return column_cursor - 1

    def write_rows_data_to_sheet(self, sheet, rows):
        """Take a list of rows containing a list of values (a value for each column)
                and write them to the passed sheet.

        to        Args:
                    sheet (openpyxl.worksheet.worksheet.Worksheet): the sheet to export to
                    rows (list): a list (rows) of list (columns) of values
        """
        for row_cursor, row in enumerate(rows, start=2):
            for column_cursor, cell_value in enumerate(row, start=1):
                sheet.cell(row=row_cursor, column=column_cursor, value=cell_value)

    def set_columns_width(self, sheet, max_width=None):
        """Set the width of each column of `sheet` to the width of the lengthy content or `max_width`.

        Args:
            sheet (WorkSheet): a sheet of a WorkBook
        """
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                if max_width is not None and len(str(cell.value)) > max_width:
                    max_length = max_width
                elif len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))

            sheet.column_dimensions[column_letter].width = max_length + 2

    def set_rows_height(self, sheet):
        """Set the height of each row of `sheet`.

        Args:
            sheet (WorkSheet): a sheet of a WorkBook
        """
        for row_index in range(0, sheet.max_row):
            sheet.row_dimensions[row_index + 1].height = 30

    def init_named_styles(self):
        for named_style in self.config.named_styles:
            try:
                self.workbook.add_named_style(named_style)
            except ValueError:
                # If the named style is already registered, do nothing
                ...

    def set_column_styles(self, sheet, max_col):
        for col in sheet.iter_cols(
            min_row=2,
            max_row=self.config.max_styled_rows,
            min_col=1,
            max_col=max_col,
        ):
            if len(col) == 0:
                continue

            style = None
            number_format = None
            column_header = self.config.get_column_header(sheet, col[0].column)
            column_config = self.config.sheets_config[sheet.title][column_header]
            if column_config is not None:
                style = column_config.get("style")
                number_format = column_config.get("number_format")
            for cell in col:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if style is not None:
                    cell.style = style

                if number_format is not None:
                    cell.number_format = number_format

    def set_version(self):
        self.workbook.custom_doc_props.append(
            IntProperty(name=VERSION_CUSTOM_PROPERTY_NAME, value=self.config.version)
        )
