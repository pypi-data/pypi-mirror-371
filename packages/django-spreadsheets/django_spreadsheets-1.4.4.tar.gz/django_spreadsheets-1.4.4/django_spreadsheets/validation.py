# Third party
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.worksheet.datavalidation import DataValidation


class ValidationMixin:
    def setup_validations(self, sheet):
        """Setup validation for given sheet based on given config and return a
        dict containing DataValidation objects to each item in config

        Args:
            sheet (Worksheet): the sheet to setup validation on

        Returns:
            list: a list of DataValidation objects with the same order as in config
        """
        validations = []
        for key, config in self.config.sheets_config[sheet.title].items():
            validation = None
            if "validation" in config:
                validation_config = config["validation"]
                if validation_config["type"] == "list":
                    validation = self.setup_list_validation(
                        validation_config,
                        name=validation_config.get("name", key),
                    )
                elif validation_config["type"] == "range":
                    validation = self.setup_range_validation(validation_config)
                elif validation_config["type"] in (
                    "date",
                    "whole",
                    "textLength",
                    "time",
                    "decimal",
                ):
                    validation = self.setup_type_validation(validation_config)

            if validation is not None:
                sheet.add_data_validation(validation)
            validations.append(validation)
        return validations

    def setup_list_validation(self, validation_config, name=""):
        """Setup a data validation on the provided choices.
        Use an inline choice list when the choices do not exceed 255 characters
        (Excel limitation), and otherwise store de choices in an hidden config
        sheet.

        Args:
            formula1 (list): an iterable describing possible values
            name (str, optional): the name used in the header row of the hidden
                                  config sheet for these choices. Defaults to "".

        Returns:
            a DataValidation object
        """
        try:
            choices = validation_config["formula1"]
            if callable(choices):
                choices = choices()
        except KeyError as e:
            raise KeyError(
                "There is no `formula1` key in your validation configuration."
            ) from e

        def format_validation_list_item(item):
            # items must not contain `,` as it is the official list's items delimiter
            return item.replace(",", self.config.validation_list_replacement_delimiter)

        prepared_choices = ",".join(
            format_validation_list_item(element)
            for element in choices
            if element is not None
        )

        if len(prepared_choices) > 255:
            # If the choices exceed 255 characters, we must store them in an
            # hidden sheet and use a range validation
            try:
                config_sheet = self.workbook[self.config.config_sheet_title]
            except KeyError:
                config_sheet = self.workbook.create_sheet(
                    title=self.config.config_sheet_title, index=99
                )
            config_sheet.sheet_state = config_sheet.SHEETSTATE_HIDDEN

            config_column_cursor = config_sheet.max_column + 1

            # Header row
            config_sheet.cell(row=1, column=config_column_cursor, value=name)

            config_row_cursor = 2
            for choice in choices:
                config_sheet.cell(
                    row=config_row_cursor, column=config_column_cursor, value=choice
                )
                config_row_cursor += 1

            validation_config[
                "choices"
            ] = f"{quote_sheetname(self.config.config_sheet_title)}!${get_column_letter(config_column_cursor)}$2:${get_column_letter(config_column_cursor)}${config_sheet.max_row}"
            return self.setup_range_validation(validation_config)

        else:
            data_validation = DataValidation(
                type="list",
                formula1=f'"{prepared_choices}"',
                allow_blank=False,
                errorStyle=validation_config.get("error_style", "warning"),
                showErrorMessage=validation_config.get("show_error_message", True),
                errorTitle=validation_config.get("error_title", "Valeur invalide"),
                error=validation_config.get(
                    "error", "Cette valeur n’est pas autorisée"
                ),
            )
            return data_validation

    def setup_range_validation(self, validation_config):
        """Setup a data validation on the possible value contained in a worksheet range.

        Args:
            range (string): a string describing a range (eg. "SheetName!A1:B4")

        Returns:
            a DataValidation object
        """
        range = validation_config["choices"]

        return DataValidation(
            type="list",
            formula1=range,
            allow_blank=False,
            errorStyle=validation_config.get("error_style", "warning"),
            showErrorMessage=validation_config.get("show_error_message", True),
            errorTitle=validation_config.get("error_title", "Valeur invalide"),
            error=validation_config.get("error", "Cette valeur n’est pas autorisée"),
        )

    def setup_type_validation(self, validation_config):
        formula1 = validation_config.get("formula1")
        if callable(formula1):
            formula1 = formula1()
        formula2 = validation_config.get("formula2")
        if callable(formula2):
            formula2 = formula2()

        return DataValidation(
            type=validation_config.get("type"),
            operator=validation_config.get("operator"),
            formula1=formula1,
            formula2=formula2,
            error=validation_config.get(
                "error", "Cette valeur n’est pas valide pour cette colonne."
            ),
            errorTitle=validation_config.get("error_title", "Valeur invalide"),
            errorStyle=validation_config.get("error_style", "warning"),
            showErrorMessage=validation_config.get("show_error_message", True),
            allow_blank=False,
        )

    def apply_validation_on_column(self, data_validations, column_index, sheet):
        """Apply validation on a column

        Args:
            data_validations (list): a list of DataValidation objects
            column_cursor (int): the 1-indexed index of the column
            sheet (Worksheet): the sheet to apply validation on
        """
        data_validation = data_validations[column_index - 1]
        if data_validation is not None:
            data_validation.add(
                f"{get_column_letter(column_index)}2:{get_column_letter(column_index)}1048576"
            )
