# Third party
from openpyxl.styles import Alignment, Font, NamedStyle, PatternFill

# Local application / specific library imports
from .conf import settings as local_settings
from .constants import DEFAULT_HEADER_STYLE_NAME


class IndexedAttributesMeta(type):
    def __new__(cls, name, bases, dct):
        index = 0
        for key, value in dct.items():
            if isinstance(value, dict):
                value["index"] = index
                index += 1
        return super().__new__(cls, name, bases, dct)


class SheetConfig(metaclass=IndexedAttributesMeta):
    """An class describing all columns configurations for a sheet:
    ```
    column_identifier = {
      "name": "name of a column",
      "required": bool,
      "required_unless": str, wether the column is required if the pointed column is empty
      "required_if": str, wether the column is required if the pointed column is filled
      "comment": str, a comment that appear on the header
      "admin_only": bool, only exported on admin mode
      "validation": {
        "type": <url|date|list|range|whole...>,
        ... see openpyxl validation
      },
      "header_style": "a named style"
      "style": "a named style"
    },
    ```

    Each attribute of the class is given a 0-based index.
    Example:
    ````
    >>> class GipsyKingsConfig(SheetConfig):
    ...    djobi = {"name: "Djobi"}
    ...    djoba = {"name": "Djoba"}

    >>> GipsyKingsEnum.djoba["index"]
    1
    ```
    """

    def __getitem__(self, key):
        return getattr(self, key)

    def items(self):
        return (
            (key, value)
            for key, value in self.__class__.__dict__.items()
            if not key.startswith("__") and not callable(value)
        )

    def keys(self):
        return [
            key
            for key, value in self.__class__.__dict__.items()
            if not key.startswith("__") and not callable(value)
        ]

    def values(self):
        return [
            value
            for key, value in self.__class__.__dict__.items()
            if not key.startswith("__") and not callable(value)
        ]

    def get_data_to_export(self):
        raise NotImplementedError


class SpreadsheetConfig:
    # A dict containing all sheets of the spreadsheet, with their names as key
    sheets_config = {}

    # Sheets that are ignored for non admin users
    admin_only_sheets = []

    # The name of the configuration sheet
    config_sheet_title = local_settings.CONFIG_SHEET_TITLE

    # Number of rows that are pre-configured with validation and formulae
    max_validation_rows = local_settings.MAX_VALIDATION_ROWS

    # Number of rows that are styled
    max_styled_rows = local_settings.MAX_STYLED_ROWS

    # List of named styles available on the whole workbook
    named_styles = [
        NamedStyle(
            name=DEFAULT_HEADER_STYLE_NAME,
            fill=PatternFill(
                start_color="000000", end_color="000000", fill_type="solid"
            ),
            font=Font(name="Calibri", color="FFFFFF", bold=True),
        ),
        NamedStyle(
            name="protected",
            fill=PatternFill("solid", fgColor="DDDDDD"),
            font=Font(name="Calibri"),
            alignment=Alignment(vertical="top", wrap_text=True),
        ),
    ]

    comment_author = local_settings.COMMENT_AUTHOR

    validation_list_replacement_delimiter = (
        local_settings.VALIDATION_LIST_REPLACEMENT_DELIMITER
    )

    version = 1

    class Meta:
        ...

    def __init__(self, include_admin_fields=False):
        self._include_admin_fields = include_admin_fields

        meta_options = self.Meta
        self.verbose_name = getattr(
            meta_options, "verbose_name", self.__class__.__name__
        )

    def __str__(self) -> str:
        return self.verbose_name

    def get_config_sheets_name(self):
        """Return a list of name of sheets valid for the current user"""
        if self._include_admin_fields:
            return self.sheets_config.keys()
        return [
            title
            for title in self.sheets_config.keys()
            if title not in self.admin_only_sheets
        ]

    def excluded_columns_indexes(self, sheet):
        """Considering the _include_admin_fields attribute, return the indexes of the columns that should not be exported.

        Args:
            sheet (Worksheet): The worksheet to consider

        Returns:
            list: a list of indexes to exclude
        """
        columns_config = self.sheets_config.get(sheet.title, {})

        return (
            [
                index
                for index, (key, config) in enumerate(columns_config.items())
                if config.get("admin_only", False)
            ]
            if not self._include_admin_fields
            else []
        )

    def get_column_header(self, sheet, index):
        excluded_columns_indexes = self.excluded_columns_indexes(sheet)
        sheet_columns_names = [
            column_header
            for index, column_header in enumerate(
                self.sheets_config.get(sheet.title, {}).keys()
            )
            if index not in excluded_columns_indexes
        ]

        try:
            return sheet_columns_names[index - 1]
        except IndexError:
            return None

    def get_column_header_with_suffix(self, column_header, column_config):
        """A "*" is appended to required columns' header"""
        if column_config.get("required", False) or column_config.get(
            "required_unless", False
        ):
            column_header = f"{column_header}*"

        return column_header

    def get_sheet_columns_configs(self, sheet):
        columns_configs = self.sheets_config.get(sheet.title, {})

        if not self._include_admin_fields:
            excluded_columns_indexes = self.excluded_columns_indexes(sheet)
            columns_configs = {
                key: value
                for index, (key, value) in enumerate(columns_configs.items())
                if index not in excluded_columns_indexes
            }

        return columns_configs

    def get_sheet_columns_configs_list(self, sheet):
        columns_configs = self.get_sheet_columns_configs(sheet)
        columns_configs = list(columns_configs.items())
        return columns_configs

    def get_validation_config(self, sheet):
        validation_config = {}
        for key, config in self.sheets_config[sheet.title].items():
            if "validation" in config:
                validation_config[key] = config["validation"]

        return validation_config

    def update_database(self, sheets):
        raise NotImplementedError
