import os
from pathlib import Path

import pandas as pd
import polars as pl
import streamlit as st
from openpyxl import load_workbook

from datasure.utils import duckdb_get_table, duckdb_save_table

# --- Get List of sheet from excel ---#


def local_excel_sheet_names(file_path: str) -> list:
    """Import an excel file and return the list of sheet names.

    PARAMS:
    -------
    file_path: str : path to the excel file
    """
    excel_file = load_workbook(file_path, read_only=True)
    sheet_names = excel_file.sheetnames

    return sheet_names


def local_read_data(filename: str, sheet_name: str | None = None) -> pd.DataFrame:
    """Import data from a file.

    PARAMS:
    -------
    filename: str : path to the file
    sheet_name: str : name of the sheet to import (only for excel files)

    Returns
    -------
    data: pd.DataFrame : imported data

    """
    # get file extension
    fileext = filename.split(".")[-1]

    # import file depending on the file extension
    if fileext == "csv":
        data = pd.read_csv(filename, encoding="utf-8")
    elif fileext in ["xlsx", "xls"]:
        data = pd.read_excel(filename, sheet_name=sheet_name, engine="openpyxl")
    elif fileext == "json":
        data = pd.read_json(filename, encoding="utf-8")
    elif fileext == "dta":
        data = pd.read_stata(filename)

    return data


# --- FORM for Adding file from local storage ---#


def local_add_form(
    project_id: str, edit_mode: bool = False, defaults: dict | None = None
) -> None:
    """Form for adding a file from local storage.

    PARAMS:
    -------
    None

    Returns
    -------
    None

    """
    mode = "edit" if edit_mode else "add"

    def valid_alias(alias: str) -> bool:
        """Validate alias for uniqueness and length."""
        if not alias:
            st.error("Alias cannot be empty")
            return False
        if len(alias) > 20:
            st.error("Alias must be a maximum of 20 characters")
            return False
        return True

    def valid_file_path(file_path: str) -> bool:
        """Validate file path for existence and type."""
        if not file_path:
            st.error("File path cannot be empty")
            return False
        if not os.path.isfile(file_path):
            st.error("File not found. Please check the file path")
            return False
        valid_extensions = ["csv", "xlsx", "xls", "json", "dta"]
        if file_path.split(".")[-1] not in valid_extensions:
            st.error("Invalid file type. Please upload a valid file type")
            return False
        return True

    # Get the path to the assets directory relative to the package
    assets_dir = Path(__file__).parent.parent / "assets"
    image_path = assets_dir / "hard-disk.png"
    st.image(str(image_path), width=100)
    st.subheader("Add File from Local Storage")

    if edit_mode:
        st.info("You are in edit mode. Please modify the file details below.")
        # load the current file details from the defaults
        default_local_file_alias = defaults.get("alias", "")
        default_local_added_file = defaults.get("filename", "")
        default_local_added_file_sheet_name = defaults.get("sheet_name", "")
    else:
        # set default values for the form inputs
        default_local_file_alias = ""
        default_local_added_file = ""
        default_local_added_file_sheet_name = ""

    local_file_alias = st.text_input(
        label="alias*",
        help="Enter a unique, short, descriptive name for the file",
        placeholder=default_local_file_alias if edit_mode else "",
        disabled=edit_mode,
    )
    if local_file_alias:
        valid_alias(local_file_alias)

    # file uploader. Limit to 1 file and allow only file types selected
    local_added_file = st.text_input(
        label="file path*",
        help="Add full file name and path. eg. C:/data/survey.dta",
        placeholder=default_local_added_file if edit_mode else "",
    )

    if local_added_file and valid_file_path(local_added_file):
        local_added_file_ext = local_added_file.split(".")[-1]

        if local_added_file_ext in ["xlsx", "xls"]:
            sheets = local_excel_sheet_names(local_added_file)

            # check if default sheetname exists in the list of sheets
            if (
                default_local_added_file_sheet_name
                and default_local_added_file_sheet_name in sheets
            ):
                # get index of the default sheet name or set to first sheet
                default_sheet_index = sheets.index(default_local_added_file_sheet_name)
            else:
                default_sheet_index = 0

            local_added_file_sheet_name = st.selectbox(
                label="Sheet Name",
                options=sheets,
                index=default_sheet_index,
            )
        else:
            local_added_file_sheet_name = ""
    else:
        local_added_file_sheet_name = ""

    # add a submit button
    local_add_btn = st.button(
        "Add File",
        type="primary",
        use_container_width=True,
        key=f"add_file_key{mode}",
        disabled=not local_added_file and not local_file_alias,
    )

    st.markdown("**required*")

    # if submit (local_add_file) button is clicked
    if local_add_btn:
        import_log = duckdb_get_table(project_id, alias="import_log", db_name="logs")

        # check that alias is unique
        if not import_log.is_empty() and (
            local_file_alias in import_log["alias"].to_list()
        ):
            st.error(
                "Alias already exists. Please choose a different alias or edit the existing one."
            )
        else:
            if edit_mode:
                # update the row in the cache file
                import_log = import_log.with_columns(
                    pl.when(pl.col("alias") == default_local_file_alias)
                    .then(pl.lit(local_added_file))
                    .otherwise(pl.col("filename"))
                    .alias("filename"),
                )

            else:
                # create a new row with the file details
                new_row = {
                    "refresh": True,
                    "load": True,
                    "alias": local_file_alias,
                    "filename": local_added_file,
                    "sheet_name": local_added_file_sheet_name,
                    "source": "local storage",
                    "server": "",
                    "username": "",
                    "form_id": "",
                    "private_key": "",
                    "save_to": "",
                    "attachments": False,
                }

                if import_log.is_empty():
                    import_log = pl.DataFrame([new_row])
                else:
                    # append the new row to the cache file
                    import_log = pl.concat(
                        [import_log, pl.DataFrame([new_row])], how="vertical"
                    )

            # save the updated cache file
            duckdb_save_table(
                project_id,
                import_log,
                alias="import_log",
                db_name="logs",
            )


# --- Load data from local storage ---#


def local_load_action(
    project_id: str, alias: str, filename: str, sheet_name: str | None
) -> None:
    """Load data from local storage.

    PARAMS:
    -------
    project_id: str : project ID
    data_index: int : index of the data to load
    alias: str : alias for the data
    filename: str : path to the file
    sheet_name: str : name of the sheet to import (if applicable)

    Returns
    -------
    None
    """
    # read data from file
    data: pl.DataFrame = local_read_data(filename, sheet_name)

    # save data to DuckDB
    duckdb_save_table(
        project_id,
        data,
        alias=alias,
        db_name="raw",
    )
