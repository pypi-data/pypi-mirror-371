# -*- coding: utf-8 -*-
"""
This function was moved from the EmberFactory with no changes so far. It could probably be improved.
However, the complexity of the parser partly comes from the file format, which has constraints related to user needs.
"""
import numpy as np
import warnings
import json
from embermaker import helpers as hlp
from embermaker.helpers import norm
from embermaker import ember as emb
from embermaker import parameters as param
from embermaker.__init__ import __version__
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
trd = emb.Transition  # Transition definitions


def read_fullflex(sheet, gp, logger):
    """
    Parse Fullflex file format = from Zommers et al 2020 (! deprecated !)
    """
    # BE data storage
    lbes = []  # list of ember instances (to be filled by reading the data part of the Excel sheet, below)
    be_risk = None
    be_group = ''
    ndata = 0  # Will be set to the number of risk levels in the file

    dstate = 'paused'  # Cannot read data until 1) allowed to by the 'Start' keyword and 2) the risk levels are set
    for row in sheet.rows:
        key = hlp.stripped(row[0].value)
        name = hlp.stripped(row[1].value)
        inda = [acell.value for acell in row[2:]]  # input data
        if key == 'RISK-INDEX':
            # Get the risk levels for which the file will provide data ('hazard' levels related to the risk levels)
            be_risk = inda
            try:
                ndata = be_risk.index('ref_to_preind')  # number of risk T(risk) levels for which there is data
                # There are two additional values in header : ref-to-pre-ind and top_value (see .xlsx file)
                dstate = 'ready'
            except ValueError:
                raise Exception("Could not find column 'ref_to_preind' in the header line.")
            del be_risk[ndata:]
            for rlev in be_risk:
                if isinstance(rlev, str):
                    raise Exception("There seems to be a missing value in the RISK-INDEX. This is not allowed")
            logger.addinfo('Read risk-index values:' + str(be_risk))
        elif key == 'START':
            dstate = 'waiting header'
            logger.addinfo('Waiting for risk levels / header')
        elif key == 'STOP':
            dstate = 'paused'
            logger.addinfo('Paused')
        elif key == 'GROUP' and dstate != 'paused':
            be_group = row[1].value
            logger.addinfo('Reading ember group: ' + str(be_group), mestype='title')
        elif key == 'HAZARD-INDICATOR' and dstate == 'waiting header':
            raise Exception("DATA was found before any risk levels / header line - cannot handle this.")
        elif key == 'HAZARD-INDICATOR' and dstate == 'ready':
            logger.addinfo('Reading data for: ' + str(name), mestype='subtitle')
            # Create an ember and add it to the list of embers:
            be = emb.Ember(name=name)
            lbes.append(be)
            be.group = be_group
            trans = None
            rhaz = float(inda[ndata])  # Reference hazard level (e.g. temperature) / pre-ind
            # Range of BE validity (do not show colours beyond that)
            # The 'fullflex' format only has an upper range so far, called 'top value' in the sheet;
            #     a range bottom was added later in the standard (formerly basic) fmt and copied here 'just in case'
            be.haz_valid = (float(gp('haz_valid_bottom', gp.get('haz_axis_bottom', 0))),
                            float(inda[ndata + 1]) + rhaz)
            be_hazl = []  # temporary storage for hazard-level data within a single ember
            be_risl = []  # temporary storage for risk-level data within a single ember
            if ndata != len(be_risk):
                return {'error': logger.addfail(
                    "Fullflex fmt: #risk levels does not appear to match #hazard levels")}

            for i, x in enumerate(inda[0:ndata]):
                if x is not None:  # so we skip risk levels with missing data for hazard level
                    try:
                        xnum = float(x)
                    except ValueError:
                        return {'error': logger.addfail(
                            "Fullflex fmt: HAZARD-INDICATOR contains non-numeric data")}
                    be_hazl.append(xnum + rhaz)
                    be_risl.append(be_risk[i])

            # Create "standard-format" Transitions (which did not exist in the original 'fullflex' format:
            # - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
            itran = -1  # Index of the current transition; -1 means we don't have one yet
            nlevs = len(be_hazl)
            trans_phase_std_names = {'p0': 'min', 'p50': 'median', 'p100': 'max'}
            # Use only standard transition names. This covers the use in Zommers et al.,
            trans_names = ['undetectable to moderate', 'moderate to high', 'high to very high']

            for ilev in range(nlevs):
                ctran = int(be_risl[ilev])  # Bottom risk level for this transition (0 = undetect., ...)
                xtran = -1  # id of transition for top of next level, if any:
                if (ilev + 1) < nlevs:
                    xtran = np.ceil(be_risl[ilev+1])
                # if itran = ctran or ctran = xtran, we are not in a new transition yet; otherwise, create one:
                if itran < ctran < xtran and ctran < len(trans_names):
                    # Create new transition:
                    trans = emb.Transition(trans_names[ctran])
                    be.trans_append(trans)
                    itran = ctran
                    logger.addinfo(trans)
                if trans is None:
                    return {'error': logger.addfail("A transition could not be defined for legacy 'fullflex' fmt'")}
                # By definition, trans_phase = risk idx - base risk idx; write that in the percentile fmt (p*):
                trans_phase = "p{:d}".format(int((be_risl[ilev] - itran)*100))
                # Upgrade that to current standard names, when possible: (that might be generalized?)
                if trans_phase in trans_phase_std_names:
                    trans_phase = trans_phase_std_names[trans_phase]
                trans.append_level(trans_phase, be_hazl[ilev])
                logger.addinfo(f"{trans_phase}: {be_hazl[ilev]}")

        elif key == 'CONFIDENCE' and dstate == 'ready':
            logger.addinfo("CONFIDENCE is not supported; "
                           "please move to the standard format if you need confidence levels.")
        elif key == 'HAZARD-INDICATOR-DEF':
            gp['haz_' + name] = inda[0]  # Those parameters are on the first sheet because they relate to data
    return lbes


def read_standard(sheet, gp, logger):
    """
    Parse standard file format = based on IPCC SRCCL sup. mat.
    """
    # BE data storage
    lbes = []  # list of ember instances (to be filled by reading the data part of the Excel sheet, below)
    be = None  # The ember currently being processed

    # List of parameters that can be used on the first sheet (= together with the data):
    # Todo: this is used beyond the original intention and becomes cluttered;
    #       rethink: include in paramdefs.md? And/or have source and figure information on a different worksheet?
    bfparams = ['project_name', 'project_source', 'project_revision_date', 'project_version',
                'source_key', 'source_title', 'source_author', 'source_editor', 'source_year',
                'source_crossref', 'source_chapter', 'source_figure_number',
                'source_figure_title', 'source_figure_caption', 'source_figure_shortname', 'source_figure_data',
                'haz_name', 'haz_name_std', 'haz_unit', 'haz_top_value', 'haz_axis_bottom', 'haz_axis_top',
                'haz_bottom_value', 'haz_valid_bottom', 'haz_valid_top',
                'haz_map_factor', 'haz_map_shift', 'be_palette',
                'leg_title', 'leg_pos', 'software_version_min']

    # All parameters have a 'global' file scope, except for those below, which can be changed in betweeen the
    # reading of ember (precaution added to block any change in haz_map_*, which would not be processed).
    perember = ['haz_valid_bottom', 'haz_valid_top', 'haz_top_value']  # (note: haz_top_value is deprecated)

    dstate = 'wait-data'  # File reading status
    # Valid file reading statuses:
    # - wait-data: default condition, first level of file analysis, not expecting a particular kind of data.
    # - reading: inside an ember; expecing a valid ember line
    # - paused: ignores anything until a "START" line is found
    started = False  # True after a first ember was read; used to check global metadata such as req. soft version.
    trans = None
    be_group = u''  # Default ember group name
    metanames = []  # List of names of ember-related metadata
    tr_explanation_row = None  # Optional row for an explanation about the transition (metadata at transition level)
    startmeta = 6

    for irow, row in enumerate(sheet.rows):
        if len(row) == 0:
            continue
        cola = hlp.stripped(row[0].value)  # Column A: may optionally contain the name of a group of embers
        colb = hlp.stripped(row[1].value)  # Column B: contains ember name, in the first line of a new ember.
        # if already reading and column D is blank or an ember name is found, an ember ended: prepare for next ember

        if dstate == 'reading' and (hlp.isempty(row[3].value) or not hlp.isempty(colb)):
            # After reading a first ember, we now expect the start of a new ember
            trans = None  # Since this is a new transition, it needs a name before it can be defined
            dstate = 'wait-data'

        if cola == 'START':
            dstate = 'wait-data'
            logger.addinfo('Waiting for ember data')
            # Check file compatibility (we should have it now because the main parameters were read)
        elif cola == 'STOP':
            dstate = 'paused'
            logger.addinfo('Paused')
        elif cola in bfparams:
            # Set parameter
            gp[cola] = colb
            if (cola not in perember) and started and cola in gp:
                logger.addfail(f"Parameter '{cola}' cannot change after reading the first ember "
                               f"(line {irow + 1} in the Excel input file)")
        elif not hlp.isempty(cola) and hlp.isempty(colb) and dstate == 'wait-data':
            # Read group name (data in first column, second column empty)
            be_group = cola
            logger.addinfo('Reading ember group: ' + str(be_group), mestype='title')
        elif colb in ['Name', 'Component', 'Name of system at risk']:
            # This line is a table header
            if len(row) > 6:
                startmeta = 6  # Index of first column potentially containing ember-related metadata.
                if norm(row[startmeta].value) == "explanation":
                    # There is an 'explanation' column for the transition
                    tr_explanation_row = startmeta
                    startmeta += 1
                # Get names of ember-related metadata - feature in development (2021); may be improved.
                metanames = [str(cell.value).strip() for cell in row[startmeta:] if cell.value]
                metanames = [name.lower().replace(' ', '_') for name in metanames if name != ""]
                for name in metanames:
                    if name not in ['remarks', 'description', 'keywords', 'longname',
                                    'inclusion_level', 'spec_references', 'justification']:
                        logger.addwarn('Unknown ember metadata name: ' + name)
        elif hlp.isempty(cola) and not hlp.isempty(colb) and dstate == 'wait-data':
            # Start new ember
            if not started:
                # This is the first ember. Check anything related to metadata before starting to read:
                fver = str(gp['software_version_min'])
                if fver > __version__:
                    logger.addwarn("The input file requires a version ({}) newer than this app ({})"
                                   .format(fver, __version__), severe=True)
                started = True
            logger.addinfo('Reading data for: ' + str(colb), mestype='subtitle')
            # Create an ember and add it to the list of embers:
            be = emb.Ember(name=colb)
            lbes.append(be)
            be.group = be_group
            # Range of BE validity (do not show colours under/above that;
            # use haz_valid_* if not None, else haz_axis* (which must at least have a default value)
            be.haz_valid = (float(gp('haz_valid_bottom', default=gp['haz_axis_bottom'])),
                            float(gp('haz_valid_top', default=gp['haz_axis_top'])))
            # Hazard metric std name
            be.haz_name_std = gp['haz_name_std'] if 'haz_name_std' in gp.keys() and gp['haz_name_std'] else None
            # Read optional metadata
            if len(row) > 6:
                nmeta = min(len(row)-6, len(metanames))  # Metadata needs a name and a value, otherwise ignored
                for icol, cell in enumerate(row[startmeta:startmeta+nmeta]):
                    be.meta[metanames[icol]] = cell.value
            dstate = 'reading'  # Now ready to read data for that ember

        if dstate == 'reading':
            # Try to read data if available, but always read transition name etc.
            trans_phase = norm(row[3].value)
            if trans_phase in trd.phases_syn:  # Old name for the phase => 'translate' to new
                trans_phase = trd.phases_syn[trans_phase]
            # Get transition name
            if not hlp.isempty(row[2].value):
                # Start of a transition
                trans_name = norm(row[2].value)
                if trans_name in trd.names_syn:  # Old name for the transition => 'translate' to new
                    trans_name = trd.names_syn[trans_name]
                conf = row[5].value  # The confidence level, if any, is given at the start of the transition.
                # (the validity of the confidence level name is checked below, when it is used)
                trans = emb.Transition(trans_name)
                if tr_explanation_row:
                    trans.meta['explanation'] = row[tr_explanation_row].value
                newtrans = True
            else:
                if trans is None:
                    return {'error': logger.addfail(f"Input file format error in line {irow + 1}. Transitions must"
                                                    "have a valid name and start with a 'min' or 'begin' value")}
                conf = hlp.stripped(row[5].value, default="")
                newtrans = False
            # Trying to read ember data, but we don't have information about the current transition:
            if (trans is None or hlp.isempty(row[3].value)) and not hlp.isempty(row[4].value):
                logger.addwarn('Input file may be badly formatted: data without transition: ' + str(row[4].value))
            trans_phper = emb.Level.phase2risk(trans_phase)
            if trans_phper is None:
                return {'error': logger.addfail(f"Input file format error: "
                                                f"unknown transition phase '{trans_phase}' in line {irow + 1}")}

            # Get and check hazard level data
            hazl = row[4].value
            if not hlp.isempty(hazl):
                logger.addinfo(f"- {trans.name}/{trans_phase} -> {str(hazl)} conf: {conf}.")
                if type(hazl) not in [float, int]:
                    try:
                        hazl = float(hazl)
                    except ValueError:
                        if hlp.norm(hazl) != "n/a":
                            logger.addwarn(f"Hazard level cannot be converted to a number: {hazl}")
                        hazl = None
                    if hazl is not None:
                        logger.addwarn(f"Hazard level is not stored as a number: {hazl}. Check Excel file?")

            # Store hazard level data as appropriate
            if not hlp.isempty(hazl):
                if newtrans:
                    # Confidence levels are attached to transitions but only known now:
                    trans.confidence = [norm(conf)]
                    be.trans_append(trans)
                elif conf:
                    trans.confidence.append(norm(conf))
                if trans_phase == 'mode':
                    # A very special case used for expert elicitation. To change later for simplification?
                    trans.mode = float(hazl)
                else:
                    # Add level to transition
                    trans.append_level(trans_phase, hazl)
    return lbes


def embers_from_xl(wb, gp=None, logger=None):
    """
    Get the ember data from an Excel workbook following a standardized format.
    There is only one current format, called "Standard" (formerly 'Basic format')
    A legacy format called "Fullflex" is still supported.
    :param wb: The Excel workbook containing the data in its first worksheet, or the specific worksheet containing data
    :param gp: Graphic Parameters which will be updated with parameters from the "data" sheet of the wb (sheet 0)
    :param logger: optional logger for reporting to the user
    :return: lbes a list of embers (lbes), or a dict containing an error message if reading was not possible
    """
    if logger is None:
        logger = hlp.Logger()
    if gp is None:
        gp = param.ParamDict(logger=logger)

    if type(wb) is Workbook:
        sht = wb.worksheets[0]
    elif type(wb) is Worksheet:
        sht = wb
    else:
        raise ValueError("embers_from_xl requires an Excel workbook or worksheet")

    # Check file format; By default, the format is "Standard" = SRCCL-like (File format, if any, is in the first 6 rows)
    ffind = [sht.cell(i, 2).value for i in range(1, 7)
             if isinstance(sht.cell(i, 1).value, str) and sht.cell(i, 1).value == "File format"]
    if len(ffind) == 1:
        ffmt = ffind[0].strip()
        if ffmt == "Basic":
            ffmt = "Standard"  # Just in case there would be such a file with the "old name" (unlikely)
    else:
        ffmt = "Standard"
    logger.addinfo("Format of the main input file: " + str(ffmt), mestype="title")

    # Specifically ignore the irrelevant warning that openpyxl cannot handle Excel's data validation feature:
    warnings.filterwarnings("ignore", module="openpyxl", message="Data Validation extension")

    if ffmt == "Fullflex":
        lbes = read_fullflex(sht, gp, logger)
    else:
        lbes = read_standard(sht, gp, logger)

    if len(lbes) == 0:
        return {'error':
                logger.addfail("No embers were found in the input file. Suspect a formatting error or incompatiblity.")}

    return lbes


def embers_from_json(jsembers: str | bytes | list | dict, logger=None):
    """
    Gets a list of ember objects from json-formatted embers data.
    :param jsembers: the data, either as json in a string or json already 'loaded' to standard python classes.
    :param logger: optional, may receive some information during the reading process
    :return: a list of Embers
    """
    if logger is None:
        logger = hlp.Logger()
    # Regardless of how jsembers were provided, we must get a list of embers (with a standard structure) from it:
    if type(jsembers) is str or type(jsembers) is bytes:
        jsembers = json.loads(jsembers)
    if type(jsembers) is dict:
        if "embers" not in jsembers:
            raise ValueError("No embers found in the data")
        jsembers: list = jsembers['embers']

    lbes = []  # Future list of Embers objects
    jsbe: dict
    for jsbe in jsembers:
        be = emb.Ember(**jsbe)
        logger.addinfo(f"Got ember {be.name}")
        for jstr in jsbe['transitions']:
            trans = emb.Transition(name=jstr['name'],  confidence=jstr['confidence'])
            be.trans_append(trans)
            jslevels: dict = jstr['levels']
            for jslv in jslevels:
                trans.append_level(jslv, jslevels[jslv])
        lbes.append(be)
    return lbes
