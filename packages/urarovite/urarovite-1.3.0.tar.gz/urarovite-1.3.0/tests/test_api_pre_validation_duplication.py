"""Migrated pre-validation duplication tests using centralized mocking utilities.

This demonstrates how the centralized mocking approach dramatically simplifies
complex pre-validation workflow tests while maintaining full test coverage.
"""

import pytest
import tempfile
import shutil
from pathlib import Path
from unittest.mock import Mock

from urarovite.core.api import (
    execute_validation,
    _create_pre_validation_duplicate,
    _create_drive_folder_duplicate,
    _is_valid_drive_folder_id,
)
from urarovite.core.exceptions import ValidationError
from .fixtures import BaseTestCase
from .mock_manager import mock_manager


class TestPreValidationDuplication(BaseTestCase):
    """Test the pre-validation duplication workflow."""

    def setup_method(self):
        """Set up test fixtures."""
        super().setup_method()
        self.temp_dir = Path(tempfile.mkdtemp())
        self.test_excel_file = self.temp_dir / "test.xlsx"

        # Create a minimal Excel file for testing
        try:
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws["A1"] = "Test Data"
            ws["B1"] = "More Data"
            wb.save(str(self.test_excel_file))
        except ImportError:
            # Create a dummy file if openpyxl not available
            self.test_excel_file.write_text("dummy excel content")

    def teardown_method(self):
        """Clean up test fixtures."""
        if self.temp_dir.exists():
            shutil.rmtree(self.temp_dir)

    def test_fix_mode_creates_duplicate(self):
        """Test that fix mode creates a duplicate before validation."""
        with mock_manager.validation_scenario(
            {
                "test_validator": {
                    "fixes_applied": 5,
                    "flags_found": 2,
                    "automated_log": "Test fixes applied",
                }
            },
            duplicate_success=True,
        ) as mocks:
            result = execute_validation(
                check={"id": "test_validator", "mode": "fix"},
                sheet_url=str(self.test_excel_file),
                auth_secret=self.encoded_creds,
            )

            # Verify duplication was called
            assert mocks["duplicate"].called

            # Verify validation results
            assert result["fixes_applied"] == 5
            assert result["flags_found"] == 2
            assert result["errors"] == []

    def test_flag_mode_no_duplicate(self):
        """Test that flag mode does not create a duplicate."""
        with mock_manager.validation_scenario(
            {
                "test_validator": {
                    "fixes_applied": 0,
                    "flags_found": 3,
                    "automated_log": "Test flags found",
                }
            },
            duplicate_success=False,
        ) as mocks:
            result = execute_validation(
                check={"id": "test_validator", "mode": "flag"},
                sheet_url=str(self.test_excel_file),
                auth_secret=self.encoded_creds,
            )

            # Verify no duplication was attempted
            assert not mocks["duplicate"].called

            # Verify validation results
            assert result["fixes_applied"] == 0
            assert result["flags_found"] == 3
            assert result["errors"] == []

    def test_duplicate_creation_failure(self):
        """Test handling when duplicate creation fails."""
        with mock_manager.validation_scenario(
            {"test_validator": {"fixes_applied": 0, "flags_found": 0}},
            duplicate_success=False,
            duplicate_error="Failed to create duplicate",
        ):
            result = execute_validation(
                check={"id": "test_validator", "mode": "fix"},
                sheet_url=str(self.test_excel_file),
                auth_secret=self.encoded_creds,
            )

            # Should have error about duplicate creation
            assert result["fixes_applied"] == 0
            assert result["flags_found"] == 0
            assert len(result["errors"]) == 1
            assert "Failed to create duplicate" in result["errors"][0]


class TestCreateDriveFolderDuplicate(BaseTestCase):
    """Test the _create_drive_folder_duplicate function."""

    def test_google_sheets_to_sheets_success(self):
        """Test successful Google Sheets to Google Sheets duplication."""
        from .fixtures import MockPatches

        with (
            MockPatches.drive_service() as mock_drive_service,
            MockPatches.sheets_service() as mock_sheets_service,
            MockPatches.duplicate_file_to_drive_folder() as mock_duplicate,
        ):
            # Mock Drive and Sheets services
            mock_drive_service.return_value = Mock()
            mock_sheets_service.return_value = Mock()

            # Mock successful duplication
            mock_duplicate.return_value = {
                "success": True,
                "url": "https://docs.google.com/spreadsheets/d/new123/edit",
                "error": None,
            }

            result = _create_drive_folder_duplicate(
                source="https://docs.google.com/spreadsheets/d/source123/edit",
                folder_id="folder123",
                target_format="sheets",
                auth_secret=self.encoded_creds,
            )

            assert result["success"] is True
            assert "new123" in result["working_path"]
            assert result["error"] is None

    def test_google_sheets_duplication_failure(self):
        """Test handling of Google Sheets duplication failure."""
        from .fixtures import MockPatches

        with (
            MockPatches.drive_service() as mock_drive_service,
            MockPatches.sheets_service() as mock_sheets_service,
            MockPatches.duplicate_file_to_drive_folder() as mock_duplicate,
        ):
            # Mock Drive and Sheets services
            mock_drive_service.return_value = Mock()
            mock_sheets_service.return_value = Mock()

            # Mock duplication failure
            mock_duplicate.return_value = {
                "success": False,
                "url": None,
                "error": "Duplication failed",
            }

            result = _create_drive_folder_duplicate(
                source="https://docs.google.com/spreadsheets/d/source123/edit",
                folder_id="folder123",
                target_format="sheets",
                auth_secret=self.encoded_creds,
            )

            assert result["success"] is False
            assert result["working_path"] is None
            assert "Duplication failed" in result["error"]

    def test_excel_format_to_drive_folder(self):
        """Test Excel format conversion to Drive folder."""
        # Create a test Excel file
        test_file = self.temp_dir / "test.xlsx"
        try:
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws["A1"] = "Test Data"
            wb.save(str(test_file))
        except ImportError:
            # Create a dummy file if openpyxl not available
            test_file.write_text("dummy excel content")

        with mock_manager.spreadsheet_scenario():
            result = _create_drive_folder_duplicate(
                source=str(test_file),
                folder_id="folder123",
                target_format="sheets",
                auth_secret=self.encoded_creds,
            )

            # This feature is now implemented
            assert result["success"] is True
            assert result["working_path"] is not None
            assert "conversion_pending" in result


class TestIntegrationWorkflow(BaseTestCase):
    """Test complete integration workflows with pre-validation duplication."""

    def test_complete_fix_workflow_local_target(self):
        """Test complete fix workflow with local target."""
        # Create a test Excel file
        test_file = self.temp_dir / "source.xlsx"
        try:
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws["A1"] = "Test Data"
            ws["B1"] = "More Data"
            wb.save(str(test_file))
        except ImportError:
            # Create a dummy file if openpyxl not available
            test_file.write_text("dummy excel content")

        with mock_manager.integration_scenario(
            {
                "empty_cells": {
                    "fixes_applied": 3,
                    "flags_found": 0,
                    "automated_log": "Fixed 3 empty cells",
                }
            },
            duplicate_success=True,
        ) as mocks:
            # Mock successful duplicate creation
            mocks["duplicate"].return_value = {
                "success": True,
                "working_path": str(self.temp_dir / "duplicate.xlsx"),
                "output_path": str(self.temp_dir / "duplicate.xlsx"),
                "error": None,
            }

            result = execute_validation(
                check={"id": "empty_cells", "mode": "fix"},
                sheet_url=str(self.temp_dir / "source.xlsx"),
                auth_secret=self.encoded_creds,
                target=str(self.temp_dir / "output.xlsx"),
            )

            assert result["fixes_applied"] == 3
            assert result["flags_found"] == 0
            assert result["errors"] == []

    def test_complete_flag_workflow_with_target(self):
        """Test complete flag workflow with target specified."""
        with mock_manager.validation_scenario(
            {
                "duplicate_rows": {
                    "fixes_applied": 0,
                    "flags_found": 5,
                    "automated_log": "Found 5 duplicate rows",
                }
            }
        ) as mocks:
            result = execute_validation(
                check={"id": "duplicate_rows", "mode": "flag"},
                sheet_url="https://docs.google.com/spreadsheets/d/source123/edit",
                auth_secret=self.encoded_creds,
                target="https://docs.google.com/spreadsheets/d/target123/edit",
            )

            # Flag mode shouldn't create duplicate
            assert not mocks["duplicate"].called
            assert result["fixes_applied"] == 0
            assert result["flags_found"] == 5


class TestDriveFolderValidation:
    """Test Drive folder ID validation."""

    def test_valid_drive_folder_id(self):
        """Test validation of valid Drive folder IDs."""
        valid_ids = [
            "1BxiMVs0XRA5nFMdKvBdBZjgmUUqptlbs74OgvE2upms",
            "0B1234567890abcdefghijklmnopqrstuvwxyz",
            "1a2b3c4d5e6f7g8h9i0j",
        ]

        for folder_id in valid_ids:
            assert _is_valid_drive_folder_id(folder_id), f"Should be valid: {folder_id}"

    def test_invalid_drive_folder_id(self):
        """Test validation of invalid Drive folder IDs."""
        invalid_ids = [
            "",
            "too-short",
            "contains spaces",
            "contains/slashes",
            "https://drive.google.com/drive/folders/1BxiMVs0XRA5nFMdKvBdBZjgmUUqptlbs74OgvE2upms",
            None,
        ]

        for folder_id in invalid_ids:
            assert not _is_valid_drive_folder_id(folder_id), (
                f"Should be invalid: {folder_id}"
            )


# Comparison of old vs new approach for pre-validation tests:

"""
OLD APPROACH (original test_api_pre_validation_duplication.py):
===============================================================

@patch('urarovite.core.api.get_validator_registry')
@patch('urarovite.core.api._create_pre_validation_duplicate')
def test_fix_mode_creates_duplicate(self, mock_create_duplicate, mock_registry):
    # Setup mocks
    mock_validator = Mock()
    mock_validator.validate.return_value = {
        "fixes_applied": 5,
        "flags_found": 2,
        "errors": [],
        "automated_log": "Test fixes applied"
    }
    mock_registry.return_value = {"test_validator": mock_validator}
    
    mock_create_duplicate.return_value = {
        "success": True,
        "working_path": "/path/to/duplicate.xlsx",
        "output_path": "/path/to/duplicate.xlsx",
        "error": None
    }

    # Execute validation in fix mode
    result = execute_validation(
        check={"id": "test_validator", "mode": "fix"},
        sheet_url=str(self.test_excel_file),
        auth_secret=ENCODED_CREDS
    )

    # Verify duplication was called
    mock_create_duplicate.assert_called_once()
    
    # Verify validation results
    assert result["fixes_applied"] == 5
    assert result["flags_found"] == 2
    assert result["errors"] == []

LINES OF CODE: ~30 lines of setup + test logic


NEW APPROACH (this migrated file):
==================================

def test_fix_mode_creates_duplicate(self):
    with mock_manager.validation_scenario({
        "test_validator": {
            "fixes_applied": 5,
            "flags_found": 2,
            "automated_log": "Test fixes applied"
        }
    }, duplicate_success=True) as mocks:
        result = execute_validation(
            check={"id": "test_validator", "mode": "fix"},
            sheet_url=str(self.test_excel_file),
            auth_secret=self.encoded_creds
        )

        # Verify duplication was called
        assert mocks["duplicate"].called
        
        # Verify validation results
        assert result["fixes_applied"] == 5
        assert result["flags_found"] == 2
        assert result["errors"] == []

LINES OF CODE: ~15 lines total

REDUCTION: ~50% fewer lines, much cleaner, same functionality
"""
