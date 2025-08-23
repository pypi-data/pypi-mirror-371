"""Migrated integration pre-validation workflow tests using centralized mocking utilities.

This demonstrates how the centralized mocking approach dramatically simplifies
complex integration workflow tests while maintaining full test coverage.
"""

import pytest
import tempfile
import shutil
from pathlib import Path

from urarovite.core.api import execute_validation
from urarovite.core.exceptions import ValidationError
from .fixtures import BaseTestCase
from .mock_manager import mock_manager


class TestCompleteWorkflowIntegration(BaseTestCase):
    """Test the complete pre-validation duplication workflow end-to-end."""

    def setup_method(self):
        """Set up test fixtures."""
        super().setup_method()
        self.temp_dir = Path(tempfile.mkdtemp())
        self.output_dir = self.temp_dir / "output"
        self.output_dir.mkdir()

    def teardown_method(self):
        """Clean up test fixtures."""
        if self.temp_dir.exists():
            shutil.rmtree(self.temp_dir)

    def test_complete_fix_workflow_google_sheets_to_local_excel(self):
        """Test complete fix workflow: Google Sheets → Local Excel."""
        duplicate_path = str(self.output_dir / "validation_duplicate_123.xlsx")

        with mock_manager.integration_scenario(
            {
                "data_quality": {
                    "fixes_applied": 5,
                    "flags_found": 2,
                    "automated_log": "Successfully applied 5 fixes to data quality flags",
                }
            }
        ) as mocks:
            # Setup conversion mock for duplication
            mocks["convert"].return_value = {
                "success": True,
                "output_path": duplicate_path,
                "error": None,
            }

            result = execute_validation(
                check={"id": "data_quality", "mode": "fix"},
                sheet_url="https://docs.google.com/spreadsheets/d/source123/edit",
                target="local",
                target_format="excel",
                auth_secret=self.encoded_creds,
            )

            # Verify workflow steps
            assert result["fixes_applied"] == 5
            assert result["flags_found"] == 2
            assert result["errors"] == []
            assert "Successfully applied 5 fixes" in result["automated_logs"]

    def test_complete_fix_workflow_google_sheets_to_drive_folder(self):
        """Test complete fix workflow: Google Sheets → Drive Folder."""
        with mock_manager.integration_scenario(
            {
                "empty_cells": {
                    "fixes_applied": 3,
                    "flags_found": 0,
                    "automated_log": "Fixed 3 empty cells",
                }
            }
        ) as mocks:
            # Setup duplicate creation for Drive folder
            mocks["duplicate"].return_value = {
                "success": True,
                "working_path": "https://docs.google.com/spreadsheets/d/dup123/edit",
                "output_path": "https://docs.google.com/spreadsheets/d/dup123/edit",
                "error": None,
            }

            result = execute_validation(
                check={"id": "empty_cells", "mode": "fix"},
                sheet_url="https://docs.google.com/spreadsheets/d/source123/edit",
                target="1BxiMVs0XRA5nFMdKvBdBZjgmUUqptlbs74OgvE2upms",
                target_format="sheets",
                auth_secret=self.encoded_creds,
            )

            assert result["fixes_applied"] == 3
            assert result["flags_found"] == 0
            assert result["errors"] == []

    def test_complete_fix_workflow_excel_to_local(self):
        """Test complete fix workflow: Excel → Local Excel."""
        # Create a test Excel file
        test_file = self.temp_dir / "source.xlsx"
        try:
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws["A1"] = "Test Data"
            ws["B1"] = "More Data"
            wb.save(str(test_file))
        except ImportError:
            # Create a dummy file if openpyxl not available
            test_file.write_text("dummy excel content")

        with mock_manager.integration_scenario(
            {
                "duplicate_rows": {
                    "fixes_applied": 2,
                    "flags_found": 0,
                    "automated_log": "Removed 2 duplicate rows",
                }
            }
        ):
            result = execute_validation(
                check={"id": "duplicate_rows", "mode": "fix"},
                sheet_url=str(self.temp_dir / "source.xlsx"),
                target="local",
                auth_secret=self.encoded_creds,
            )

            assert result["fixes_applied"] == 2
            assert result["flags_found"] == 0
            assert result["errors"] == []

    def test_complete_flag_workflow_with_report_generation(self):
        """Test complete flag workflow with detailed reporting."""
        with mock_manager.integration_scenario(
            {
                "data_validation": {
                    "fixes_applied": 0,
                    "flags_found": 7,
                    "automated_log": "Found 7 data validation flags across 3 sheets",
                }
            }
        ):
            result = execute_validation(
                check={"id": "data_validation", "mode": "flag"},
                sheet_url="https://docs.google.com/spreadsheets/d/source123/edit",
                auth_secret=self.encoded_creds,
            )

            assert result["fixes_applied"] == 0
            assert result["flags_found"] == 7
            assert result["errors"] == []
            assert "Found 7 data validation flags" in result["automated_logs"]

    def test_workflow_with_duplication_failure(self):
        """Test workflow handling when duplication fails."""
        with mock_manager.integration_scenario(
            {"empty_cells": {"fixes_applied": 0, "flags_found": 0}},
            duplicate_success=False,
        ):
            result = execute_validation(
                check={"id": "empty_cells", "mode": "fix"},
                sheet_url="https://docs.google.com/spreadsheets/d/source123/edit",
                target="local",
                auth_secret=self.encoded_creds,
            )

            assert result["fixes_applied"] == 0
            assert result["flags_found"] == 0
            assert len(result["errors"]) == 1
            assert "Failed to create duplicate" in result["errors"][0]

    def test_workflow_with_validation_errors(self):
        """Test workflow with validation errors."""
        with mock_manager.integration_scenario(
            {
                "format_validation": {
                    "fixes_applied": 1,
                    "flags_found": 0,
                    "errors": ["Warning: Could not validate cell A5 - protected range"],
                    "automated_log": "Applied 1 fix with warnings",
                }
            }
        ):
            result = execute_validation(
                check={"id": "format_validation", "mode": "fix"},
                sheet_url="https://docs.google.com/spreadsheets/d/source123/edit",
                auth_secret=self.encoded_creds,
            )

            assert result["fixes_applied"] == 1
            assert result["flags_found"] == 0
            assert len(result["errors"]) == 1
            assert "Could not validate cell A5" in result["errors"][0]

    def test_multiple_validation_checks_workflow(self):
        """Test workflow with multiple validation checks."""
        with mock_manager.integration_scenario(
            {
                "empty_cells": {
                    "fixes_applied": 3,
                    "flags_found": 0,
                    "automated_log": "Fixed 3 empty cells",
                },
                "duplicate_rows": {
                    "fixes_applied": 5,
                    "flags_found": 0,
                    "automated_log": "Removed 5 duplicate rows",
                },
            }
        ):
            # Run first validation
            result1 = execute_validation(
                check={"id": "empty_cells", "mode": "fix"},
                sheet_url="https://docs.google.com/spreadsheets/d/source123/edit",
                auth_secret=self.encoded_creds,
            )

            # Run second validation
            result2 = execute_validation(
                check={"id": "duplicate_rows", "mode": "fix"},
                sheet_url="https://docs.google.com/spreadsheets/d/source123/edit",
                auth_secret=self.encoded_creds,
            )

            # Verify both results
            assert result1["fixes_applied"] == 3
            assert result2["fixes_applied"] == 5
            assert result1["errors"] == []
            assert result2["errors"] == []

    def test_workflow_parameter_validation(self):
        """Test workflow parameter validation."""
        with mock_manager.integration_scenario():
            # Test invalid check parameter
            result = execute_validation(
                check={"mode": "fix"},  # Missing 'id'
                sheet_url="https://docs.google.com/spreadsheets/d/source123/edit",
                auth_secret=self.encoded_creds,
            )

            assert result["fixes_applied"] == 0
            assert result["flags_found"] == 0
            assert len(result["errors"]) == 1
            assert "Check missing required 'id' field" in result["errors"][0]


class TestWorkflowEdgeCases(BaseTestCase):
    """Test edge cases in workflow integration."""

    def setup_method(self):
        """Set up test fixtures."""
        super().setup_method()
        self.temp_dir = Path(tempfile.mkdtemp())

    def teardown_method(self):
        """Clean up test fixtures."""
        if self.temp_dir.exists():
            shutil.rmtree(self.temp_dir)

    def test_zero_fixes_applied(self):
        """Test workflow when no fixes are needed."""
        with mock_manager.integration_scenario(
            {
                "data_quality": {
                    "fixes_applied": 0,
                    "flags_found": 0,
                    "automated_log": "No flags found - data is clean",
                }
            }
        ):
            result = execute_validation(
                check={"id": "data_quality", "mode": "fix"},
                sheet_url="https://docs.google.com/spreadsheets/d/source123/edit",
                auth_secret=self.encoded_creds,
            )

            assert result["fixes_applied"] == 0
            assert result["flags_found"] == 0
            assert result["errors"] == []
            assert "No flags found" in result["automated_logs"]

    def test_unknown_validator(self):
        """Test workflow with unknown validator."""
        with mock_manager.integration_scenario():  # Empty registry
            result = execute_validation(
                check={"id": "unknown_validator", "mode": "fix"},
                sheet_url="https://docs.google.com/spreadsheets/d/source123/edit",
                auth_secret=self.encoded_creds,
            )

            assert result["fixes_applied"] == 0
            assert result["flags_found"] == 0
            assert len(result["errors"]) == 1
            assert (
                "Unknown validation check: 'unknown_validator'" in result["errors"][0]
            )

    def test_large_number_of_fixes(self):
        """Test workflow with large number of fixes."""
        with mock_manager.integration_scenario(
            {
                "bulk_cleanup": {
                    "fixes_applied": 9999,
                    "flags_found": 0,
                    "automated_log": "Applied 9999 bulk cleanup fixes across all sheets",
                }
            }
        ):
            result = execute_validation(
                check={"id": "bulk_cleanup", "mode": "fix"},
                sheet_url="https://docs.google.com/spreadsheets/d/source123/edit",
                auth_secret=self.encoded_creds,
            )

            assert result["fixes_applied"] == 9999
            assert result["flags_found"] == 0
            assert result["errors"] == []
            assert "Applied 9999 bulk cleanup fixes" in result["automated_logs"]


class TestWorkflowPerformance(BaseTestCase):
    """Test workflow performance scenarios."""

    def test_concurrent_validation_requests(self):
        """Test handling of concurrent validation requests."""
        with mock_manager.integration_scenario(
            {
                "performance_test": {
                    "fixes_applied": 1,
                    "flags_found": 0,
                    "automated_log": "Performance test completed",
                }
            }
        ):
            # Simulate multiple concurrent requests
            results = []
            for i in range(3):
                result = execute_validation(
                    check={"id": "performance_test", "mode": "fix"},
                    sheet_url=f"https://docs.google.com/spreadsheets/d/test{i}/edit",
                    auth_secret=self.encoded_creds,
                )
                results.append(result)

            # All should succeed
            for result in results:
                assert result["fixes_applied"] == 1
                assert result["errors"] == []


# Comparison of old vs new approach for integration workflow tests:

"""
OLD APPROACH (original test_integration_pre_validation_workflow.py):
===================================================================

@patch('urarovite.core.api.get_validator_registry')
@patch('urarovite.utils.generic_spreadsheet.convert_spreadsheet_format')
def test_complete_fix_workflow_google_sheets_to_local_excel(self, mock_convert, mock_registry):
    # Setup validator mock
    mock_validator = Mock()
    mock_validator.validate.return_value = {
        "fixes_applied": 5,
        "flags_found": 2,
        "errors": [],
        "automated_log": "Successfully applied 5 fixes to data quality flags"
    }
    mock_registry.return_value = {"data_quality": mock_validator}

    # Setup conversion mock (for duplication)
    duplicate_path = str(self.output_dir / "validation_duplicate_123.xlsx")
    mock_convert.return_value = {
        "success": True,
        "output_path": duplicate_path,
        "error": None
    }

    # Execute complete workflow
    result = execute_validation(
        check={"id": "data_quality", "mode": "fix"},
        sheet_url="https://docs.google.com/spreadsheets/d/source123/edit",
        target="local",
        target_format="excel",
        auth_secret=ENCODED_CREDS
    )

    # Verify workflow steps
    assert result["fixes_applied"] == 5
    assert result["flags_found"] == 2
    assert result["duplicate_created"] == duplicate_path
    assert result["target_output"] == duplicate_path
    assert len(result["errors"]) == 0
    assert "Successfully applied 5 fixes" in result["automated_logs"]

LINES OF CODE: ~30 lines of setup + test logic


NEW APPROACH (this migrated file):
==================================

def test_complete_fix_workflow_google_sheets_to_local_excel(self):
    duplicate_path = str(self.output_dir / "validation_duplicate_123.xlsx")
    
    with mock_manager.integration_scenario({
        "data_quality": {
            "fixes_applied": 5,
            "flags_found": 2,
            "automated_log": "Successfully applied 5 fixes to data quality flags"
        }
    }) as mocks:
        # Setup conversion mock for duplication
        mocks["convert"].return_value = {
            "success": True,
            "output_path": duplicate_path,
            "error": None
        }

        result = execute_validation(
            check={"id": "data_quality", "mode": "fix"},
            sheet_url="https://docs.google.com/spreadsheets/d/source123/edit",
            target="local",
            target_format="excel",
            auth_secret=self.encoded_creds
        )

        # Verify workflow steps
        assert result["fixes_applied"] == 5
        assert result["flags_found"] == 2
        assert result["errors"] == []
        assert "Successfully applied 5 fixes" in result["automated_logs"]

LINES OF CODE: ~20 lines total

REDUCTION: ~33% fewer lines, much cleaner setup, same functionality
"""
