"""Excel implementation of the spreadsheet interface.

This module provides Excel-specific implementation of the SpreadsheetInterface
using openpyxl for reading and writing Excel files.
"""

import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from urarovite.core.spreadsheet import (
    SpreadsheetInterface,
    SpreadsheetMetadata,
    SheetData,
)
from urarovite.core.exceptions import ValidationError

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils import column_index_from_string

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExcelSpreadsheet(SpreadsheetInterface):
    """Excel implementation of SpreadsheetInterface using openpyxl."""

    def __init__(
        self,
        file_path: Path,
        read_only: bool = False,
        preserve_formulas: bool = True,
        create_new: bool = False,
    ) -> None:
        """Initialize Excel spreadsheet.

        Args:
            file_path: Path to the Excel file
            read_only: Whether to open in read-only mode
            preserve_formulas: Whether to preserve formulas (False = values only)
            create_new: Whether to create a new Excel file if it doesn't exist

        Raises:
            ValidationError: If unable to initialize Excel file access
        """
        if not OPENPYXL_AVAILABLE:
            raise ValidationError(
                "openpyxl is required for Excel support. "
                "Install with: pip install openpyxl"
            )

        self.file_path = Path(file_path)
        self.read_only = read_only
        self.preserve_formulas = preserve_formulas

        # Handle file existence and creation
        if not self.file_path.exists():
            if create_new:
                # Create a new Excel file
                try:
                    # Ensure the directory exists
                    self.file_path.parent.mkdir(parents=True, exist_ok=True)

                    # Create new workbook
                    self.workbook = Workbook()

                    # Save the new file
                    self.workbook.save(str(self.file_path))

                except Exception as e:
                    raise ValidationError(f"Failed to create new Excel file: {str(e)}")
            else:
                raise ValidationError(f"Excel file not found: {self.file_path}")
        else:
            # Load existing file
            try:
                # Load workbook with formula preservation based on parameter
                self.workbook = load_workbook(
                    filename=str(self.file_path),
                    read_only=read_only,
                    data_only=not preserve_formulas,  # False = keep formulas, True = get values
                )
            except Exception as e:
                raise ValidationError(f"Failed to load Excel file: {str(e)}")

        self._metadata: Optional[SpreadsheetMetadata] = None

    def get_metadata(self) -> SpreadsheetMetadata:
        """Get Excel file metadata."""
        if self._metadata is None:
            try:
                sheet_names = list(self.workbook.sheetnames)
                title = self.file_path.stem  # Use filename without extension as title

                self._metadata = SpreadsheetMetadata(
                    spreadsheet_id=str(self.file_path),
                    title=title,
                    sheet_names=sheet_names,
                    spreadsheet_type="excel",
                    file_path=self.file_path,
                )

            except Exception as e:
                raise ValidationError(f"Failed to get Excel metadata: {str(e)}")

        return self._metadata

    def get_sheet_data(
        self, sheet_name: Optional[str] = None, range_name: Optional[str] = None
    ) -> SheetData:
        """Get data from Excel sheet."""
        try:
            # Get sheet name if not provided
            if not sheet_name:
                metadata = self.get_metadata()
                if not metadata.sheet_names:
                    raise ValidationError("No sheets found in Excel file")
                sheet_name = metadata.sheet_names[0]

            # Get worksheet
            if sheet_name not in self.workbook.sheetnames:
                raise ValidationError(f"Sheet '{sheet_name}' not found in Excel file")

            worksheet = self.workbook[sheet_name]

            # Parse range if provided
            if range_name:
                start_row, start_col, end_row, end_col = self._parse_range(range_name)
            else:
                # Use entire used range
                start_row, start_col = 1, 1
                end_row = worksheet.max_row
                end_col = worksheet.max_column

            # Extract data using OPTIMIZED bulk reading
            values = []
            actual_rows = 0
            actual_cols = 0

            if range_name:
                # Use range-based bulk reading for specific ranges
                try:
                    from openpyxl.utils import get_column_letter

                    end_col_letter = get_column_letter(end_col)
                    cell_range = worksheet[
                        f"{get_column_letter(start_col)}{start_row}:{end_col_letter}{end_row}"
                    ]

                    for row_idx, row in enumerate(cell_range):
                        row_values = []
                        row_has_data = False

                        for col_idx, cell in enumerate(row):
                            value = cell.value

                            # Convert None to empty string for consistency with Google Sheets
                            if value is None:
                                value = ""

                            row_values.append(value)

                            if value != "" and value is not None:
                                row_has_data = True
                                actual_cols = max(actual_cols, col_idx + 1)

                        values.append(row_values)

                        if row_has_data:
                            actual_rows = row_idx + 1

                except Exception:
                    # Fallback to iter_rows if range reading fails
                    for row in worksheet.iter_rows(
                        min_row=start_row,
                        max_row=end_row,
                        min_col=start_col,
                        max_col=end_col,
                        values_only=True,
                    ):
                        row_values = []
                        row_has_data = False

                        for col_idx, value in enumerate(row):
                            # Convert None to empty string for consistency with Google Sheets
                            if value is None:
                                value = ""

                            row_values.append(value)

                            if value != "" and value is not None:
                                row_has_data = True
                                actual_cols = max(actual_cols, col_idx + 1)

                        values.append(row_values)

                        if row_has_data:
                            actual_rows = len(values)
            else:
                # Use ultra-fast bulk reading for entire sheet
                if self.read_only:
                    # For read-only mode, use the fastest method
                    try:
                        # Method 1: Try worksheet.values (fastest)
                        sheet_values = list(worksheet.values)

                        for row_idx, row in enumerate(sheet_values):
                            if row is None:
                                continue

                            row_values = []
                            row_has_data = False

                            for col_idx, value in enumerate(row):
                                # Convert None to empty string for consistency with Google Sheets
                                if value is None:
                                    value = ""

                                row_values.append(value)

                                if value != "" and value is not None:
                                    row_has_data = True
                                    actual_cols = max(actual_cols, col_idx + 1)

                            values.append(row_values)

                            if row_has_data:
                                actual_rows = row_idx + 1

                    except Exception:
                        # Fallback to iter_rows
                        for row in worksheet.iter_rows(values_only=True):
                            if row is None:
                                continue

                            row_values = []
                            row_has_data = False

                            for col_idx, value in enumerate(row):
                                # Convert None to empty string for consistency with Google Sheets
                                if value is None:
                                    value = ""

                                row_values.append(value)

                                if value != "" and value is not None:
                                    row_has_data = True
                                    actual_cols = max(actual_cols, col_idx + 1)

                            values.append(row_values)

                            if row_has_data:
                                actual_rows = len(values)
                else:
                    # For write mode, use iter_rows for better compatibility
                    for row in worksheet.iter_rows(
                        min_row=start_row,
                        max_row=end_row,
                        min_col=start_col,
                        max_col=end_col,
                        values_only=True,
                    ):
                        row_values = []
                        row_has_data = False

                        for col_idx, value in enumerate(row):
                            # Convert None to empty string for consistency with Google Sheets
                            if value is None:
                                value = ""

                            row_values.append(value)

                            if value != "" and value is not None:
                                row_has_data = True
                                actual_cols = max(actual_cols, col_idx + 1)

                        values.append(row_values)

                        if row_has_data:
                            actual_rows = len(values)

            return SheetData(
                values=values, sheet_name=sheet_name, rows=actual_rows, cols=actual_cols
            )

        except Exception as e:
            if isinstance(e, ValidationError):
                raise
            raise ValidationError(f"Failed to get Excel sheet data: {str(e)}")

    def get_sheet_data_with_hyperlinks(
        self,
        sheet_name: Optional[str] = None,
        range_name: Optional[str] = None,
    ) -> SheetData:
        """Get data from Excel, resolving hyperlinks."""
        try:
            if not sheet_name:
                metadata = self.get_metadata()
                if not metadata.sheet_names:
                    raise ValidationError("No sheets found in Excel file")
                sheet_name = metadata.sheet_names[0]

            if sheet_name not in self.workbook.sheetnames:
                raise ValidationError(f"Sheet '{sheet_name}' not found.")

            worksheet = self.workbook[sheet_name]
            values = []
            for row in worksheet.iter_rows(values_only=False):
                row_values = []
                for cell in row:
                    if cell.hyperlink and cell.hyperlink.target:
                        row_values.append(cell.hyperlink.target)
                    else:
                        row_values.append(cell.value if cell.value is not None else "")
                values.append(row_values)

            rows = len(values)
            cols = max(len(row) for row in values) if values else 0

            return SheetData(values=values, sheet_name=sheet_name, rows=rows, cols=cols)

        except Exception as e:
            if isinstance(e, ValidationError):
                raise
            raise ValidationError(f"Failed to get Excel hyperlinks: {str(e)}")

    def update_sheet_data(
        self,
        sheet_name: str,
        values: List[List[Any]],
        start_row: int = 1,
        start_col: int = 1,
        range_name: Optional[str] = None,
    ) -> None:
        """Update data in Excel sheet."""
        if self.read_only:
            raise ValidationError("Cannot update data in read-only mode")

        try:
            # Get worksheet
            if sheet_name not in self.workbook.sheetnames:
                raise ValidationError(f"Sheet '{sheet_name}' not found in Excel file")

            worksheet = self.workbook[sheet_name]

            # Parse range if provided
            if range_name:
                start_row, start_col, _, _ = self._parse_range(range_name)

            # Write data
            for row_idx, row_data in enumerate(values):
                for col_idx, cell_value in enumerate(row_data):
                    cell = worksheet.cell(
                        row=start_row + row_idx, column=start_col + col_idx
                    )
                    cell.value = cell_value

        except Exception as e:
            if isinstance(e, ValidationError):
                raise
            raise ValidationError(f"Failed to update Excel sheet data: {str(e)}")

    def update_sheet_formulas(
        self, sheet_name: str, formulas: Dict[str, str], preserve_values: bool = True
    ) -> None:
        """Update formulas in Excel sheet.

        Args:
            sheet_name: Name of the sheet to update
            formulas: Dict mapping cell coordinates to formulas
            preserve_values: Whether to preserve existing values for non-formula cells
        """
        if self.read_only:
            raise ValidationError("Cannot update formulas in read-only mode")

        if not self.preserve_formulas:
            # If workbook was loaded with data_only=True, we can't set formulas
            import logging

            logging.warning(
                f"Cannot set formulas in Excel file loaded with preserve_formulas=False. "
                f"Formulas will be lost: {list(formulas.keys())}"
            )
            return

        try:
            # Get worksheet
            if sheet_name not in self.workbook.sheetnames:
                raise ValidationError(f"Sheet '{sheet_name}' not found in Excel file")

            worksheet = self.workbook[sheet_name]

            # Update formulas
            for cell_ref, formula in formulas.items():
                try:
                    # Ensure formula starts with =
                    if not formula.startswith("="):
                        formula = "=" + formula

                    # Set the formula
                    worksheet[cell_ref] = formula

                except Exception as e:
                    import logging

                    logging.warning(
                        f"Failed to set formula in cell {cell_ref}: {formula}. Error: {e}"
                    )

        except Exception as e:
            if isinstance(e, ValidationError):
                raise
            raise ValidationError(f"Failed to update Excel formulas: {str(e)}")

    def update_sheet_properties(
        self, sheet_name: str, new_name: Optional[str] = None, **properties: Any
    ) -> None:
        """Update Excel sheet properties."""
        if self.read_only:
            raise ValidationError("Cannot update properties in read-only mode")

        try:
            if sheet_name not in self.workbook.sheetnames:
                raise ValidationError(f"Sheet '{sheet_name}' not found in Excel file")

            worksheet = self.workbook[sheet_name]

            if new_name:
                # Validate new name
                if new_name in self.workbook.sheetnames and new_name != sheet_name:
                    raise ValidationError(f"Sheet name '{new_name}' already exists")

                # Check Excel naming constraints and truncate if necessary
                if len(new_name) > 31:
                    # Truncate to 31 characters, preserving as much of the original as possible
                    truncated_name = new_name[:31]
                    import logging
                    logging.info(f"Truncated sheet name '{new_name}' to '{truncated_name}' for Excel compatibility")
                    new_name = truncated_name

                illegal_chars = ["\\", "/", "?", "*", "[", "]"]
                if any(char in new_name for char in illegal_chars):
                    raise ValidationError(
                        f"Sheet name contains illegal characters: {illegal_chars}"
                    )

                worksheet.title = new_name

                # Invalidate cached metadata
                self._metadata = None

            # Handle other properties if needed (color, visibility, etc.)
            for prop_name, prop_value in properties.items():
                if hasattr(worksheet, prop_name):
                    setattr(worksheet, prop_name, prop_value)

        except Exception as e:
            if isinstance(e, ValidationError):
                raise
            raise ValidationError(f"Failed to update Excel sheet properties: {str(e)}")

    def create_sheet(self, sheet_name: str) -> None:
        """Create a new sheet in Excel file."""
        if self.read_only:
            raise ValidationError("Cannot create sheet in read-only mode")

        try:
            if sheet_name in self.workbook.sheetnames:
                raise ValidationError(f"Sheet '{sheet_name}' already exists")

            # Check Excel naming constraints and truncate if necessary
            if len(sheet_name) > 31:
                # Truncate to 31 characters, preserving as much of the original as possible
                truncated_name = sheet_name[:31]
                import logging
                logging.info(f"Truncated sheet name '{sheet_name}' to '{truncated_name}' for Excel compatibility")
                sheet_name = truncated_name

            illegal_chars = ["\\", "/", "?", "*", "[", "]"]
            if any(char in sheet_name for char in illegal_chars):
                raise ValidationError(
                    f"Sheet name contains illegal characters: {illegal_chars}"
                )

            self.workbook.create_sheet(title=sheet_name)

            # Invalidate cached metadata
            self._metadata = None

        except Exception as e:
            if isinstance(e, ValidationError):
                raise
            raise ValidationError(f"Failed to create Excel sheet: {str(e)}")

    def delete_sheet(self, sheet_name: str) -> None:
        """Delete a sheet from Excel file."""
        if self.read_only:
            raise ValidationError("Cannot delete sheet in read-only mode")

        try:
            if sheet_name not in self.workbook.sheetnames:
                raise ValidationError(f"Sheet '{sheet_name}' not found")

            if len(self.workbook.sheetnames) <= 1:
                raise ValidationError("Cannot delete the last remaining sheet")

            worksheet = self.workbook[sheet_name]
            self.workbook.remove(worksheet)

            # Invalidate cached metadata
            self._metadata = None

        except Exception as e:
            if isinstance(e, ValidationError):
                raise
            raise ValidationError(f"Failed to delete Excel sheet: {str(e)}")

    def save(self) -> None:
        """Save changes to Excel file."""
        if self.read_only:
            raise ValidationError("Cannot save in read-only mode")

        try:
            self.workbook.save(filename=str(self.file_path))
        except Exception as e:
            raise ValidationError(f"Failed to save Excel file: {str(e)}")

    def close(self) -> None:
        """Close Excel workbook and clean up resources."""
        try:
            if hasattr(self.workbook, "close"):
                self.workbook.close()
            self._metadata = None
        except Exception as e:
            # Don't raise error on close, just log
            import logging

            logging.warning(f"Warning during Excel file close: {str(e)}")

    def get_formulas(
        self, sheet_name: Optional[str] = None
    ) -> Dict[str, Dict[str, str]]:
        """Get formulas from Excel sheets.

        Args:
            sheet_name: Name of specific sheet (optional, gets all sheets if None)

        Returns:
            Dict mapping sheet names to dict of cell coordinates to formulas
        """
        try:
            # Load workbook without data_only to get formulas
            formula_workbook = load_workbook(
                filename=str(self.file_path),
                read_only=True,
                data_only=False,  # Get formulas instead of calculated values
            )

            formulas = {}

            # Determine which sheets to process
            if sheet_name:
                if sheet_name not in formula_workbook.sheetnames:
                    raise ValidationError(
                        f"Sheet '{sheet_name}' not found in Excel file"
                    )
                sheets_to_process = [sheet_name]
            else:
                sheets_to_process = formula_workbook.sheetnames

            # Extract formulas from each sheet
            for sheet in sheets_to_process:
                worksheet = formula_workbook[sheet]
                sheet_formulas = {}

                # Iterate through all cells in the used range
                for row in worksheet.iter_rows():
                    for cell in row:
                        if (
                            cell.value
                            and isinstance(cell.value, str)
                            and cell.value.startswith("=")
                        ):
                            sheet_formulas[cell.coordinate] = cell.value

                if sheet_formulas:  # Only add sheets that have formulas
                    formulas[sheet] = sheet_formulas

            formula_workbook.close()
            return formulas

        except Exception as e:
            raise ValidationError(f"Failed to read formulas from Excel file: {str(e)}")

    def _parse_range(self, range_name: str) -> Tuple[int, int, int, int]:
        """Parse A1 notation range into row/column indices.

        Args:
            range_name: A1 notation range (e.g., 'A1:Z100', 'B2:D10')

        Returns:
            Tuple of (start_row, start_col, end_row, end_col) - all 1-based

        Raises:
            ValidationError: If range format is invalid
        """
        try:
            # Remove sheet name if present
            if "!" in range_name:
                _, range_part = range_name.split("!", 1)
            else:
                range_part = range_name

            range_part = range_part.strip()

            if ":" in range_part:
                start_cell, end_cell = range_part.split(":", 1)
            else:
                # Single cell
                start_cell = end_cell = range_part

            # Parse start cell
            start_row, start_col = self._parse_cell_reference(start_cell.strip())

            # Parse end cell
            end_row, end_col = self._parse_cell_reference(end_cell.strip())

            return start_row, start_col, end_row, end_col

        except Exception as e:
            raise ValidationError(f"Invalid range format '{range_name}': {str(e)}")

    def _parse_cell_reference(self, cell_ref: str) -> Tuple[int, int]:
        """Parse cell reference like 'A1' into row/column indices.

        Args:
            cell_ref: Cell reference (e.g., 'A1', 'Z100')

        Returns:
            Tuple of (row, col) - both 1-based

        Raises:
            ValidationError: If cell reference format is invalid
        """
        try:
            # Use regex to split column letters and row numbers
            match = re.match(r"^([A-Z]+)(\d+)$", cell_ref.upper())
            if not match:
                raise ValidationError(f"Invalid cell reference: {cell_ref}")

            col_letters, row_str = match.groups()

            # Convert column letters to index
            col_idx = column_index_from_string(col_letters)
            row_idx = int(row_str)

            return row_idx, col_idx

        except Exception as e:
            raise ValidationError(f"Invalid cell reference '{cell_ref}': {str(e)}")
