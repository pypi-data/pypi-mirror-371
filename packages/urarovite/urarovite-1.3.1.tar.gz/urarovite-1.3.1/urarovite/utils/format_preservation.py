"""Visual formatting preservation utilities.

This module provides comprehensive visual formatting preservation when converting
between Excel and Google Sheets, including fonts, colors, borders, and alignment.
"""

import logging
from typing import Any, Dict, List, Optional, Union, Tuple
from dataclasses import dataclass
from pathlib import Path

logger = logging.getLogger(__name__)

try:
    from openpyxl.styles import Font, Fill, Border, Alignment, PatternFill
    from openpyxl.styles.colors import Color

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


@dataclass
class CellFormat:
    """Represents formatting for a single cell."""

    # Font properties
    font_name: Optional[str] = None
    font_size: Optional[float] = None
    font_bold: Optional[bool] = None
    font_italic: Optional[bool] = None
    font_underline: Optional[bool] = None
    font_color: Optional[str] = None  # Hex color like "#FF0000"

    # Background/Fill properties
    background_color: Optional[str] = None  # Hex color like "#FFFF00"

    # Border properties
    border_top: Optional[str] = None
    border_bottom: Optional[str] = None
    border_left: Optional[str] = None
    border_right: Optional[str] = None
    border_color: Optional[str] = None

    # Alignment properties
    horizontal_align: Optional[str] = None  # "left", "center", "right"
    vertical_align: Optional[str] = None  # "top", "middle", "bottom"
    text_wrap: Optional[bool] = None

    # Number format
    number_format: Optional[str] = None  # "0.00", "mm/dd/yyyy", etc.


class FormatExtractor:
    """Extracts formatting information from spreadsheets."""

    def extract_excel_formatting(
        self, workbook: Any, sheet_name: str
    ) -> Dict[str, CellFormat]:
        """Extract formatting from Excel worksheet.

        Args:
            workbook: openpyxl workbook object
            sheet_name: Name of the worksheet

        Returns:
            Dict mapping cell coordinates (A1, B2, etc.) to CellFormat objects
        """
        if not OPENPYXL_AVAILABLE:
            logger.warning("openpyxl not available - cannot extract Excel formatting")
            return {}

        try:
            worksheet = workbook[sheet_name]
            cell_formats = {}

            # Iterate through all cells that have data or formatting
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value is not None or self._has_formatting(cell):
                        cell_ref = cell.coordinate
                        cell_formats[cell_ref] = self._extract_excel_cell_format(cell)

            logger.info(
                f"Extracted formatting for {len(cell_formats)} cells from {sheet_name}"
            )
            return cell_formats

        except Exception as e:
            logger.error(f"Failed to extract Excel formatting from {sheet_name}: {e}")
            return {}

    def _has_formatting(self, cell: Any) -> bool:
        """Check if a cell has any formatting applied."""
        try:
            # Check for non-default font
            if (
                cell.font.name != "Calibri"
                or cell.font.size != 11
                or cell.font.bold
                or cell.font.italic
            ):
                return True

            # Check for background color (be more strict about what counts as formatting)
            if (
                hasattr(cell.fill, "fgColor")
                and cell.fill.fgColor
                and cell.fill.fgColor.rgb
                and cell.fill.fgColor.rgb != "00000000"  # Not black/default
                and cell.fill.fgColor.rgb != "FFFFFFFF"  # Not white/default
                and cell.fill.fgColor.rgb != "FF000000"  # Not default black
                and len(cell.fill.fgColor.rgb) == 8  # Valid RGB format
                and cell.fill.fill_type  # Has actual fill type
                and cell.fill.fill_type != "none"
            ):  # Not "no fill"
                return True

            # Check for borders
            if (
                cell.border.top.style
                or cell.border.bottom.style
                or cell.border.left.style
                or cell.border.right.style
            ):
                return True

            # Check for alignment
            if (
                cell.alignment.horizontal != "general"
                or cell.alignment.vertical != "bottom"
            ):
                return True

            return False

        except Exception:
            return False

    def _extract_excel_cell_format(self, cell: Any) -> CellFormat:
        """Extract formatting from a single Excel cell."""
        try:
            format_obj = CellFormat()

            # Font properties
            if cell.font:
                format_obj.font_name = cell.font.name
                format_obj.font_size = cell.font.size
                format_obj.font_bold = cell.font.bold
                format_obj.font_italic = cell.font.italic
                format_obj.font_underline = cell.font.underline is not None

                # Font color
                if hasattr(cell.font, "color") and cell.font.color:
                    format_obj.font_color = self._excel_color_to_hex(cell.font.color)

            # Background color (be selective about which colors to preserve)
            if cell.fill and hasattr(cell.fill, "fgColor"):
                bg_color = self._excel_color_to_hex(cell.fill.fgColor)
                # Skip problematic colors (black, white, or default colors)
                if (
                    bg_color
                    and bg_color.lower()
                    not in ["#000000", "#ffffff", "#ff000000", "#ffffffff"]
                    and cell.fill.fill_type
                    and cell.fill.fill_type != "none"
                ):
                    format_obj.background_color = bg_color

            # Borders
            if cell.border:
                format_obj.border_top = cell.border.top.style
                format_obj.border_bottom = cell.border.bottom.style
                format_obj.border_left = cell.border.left.style
                format_obj.border_right = cell.border.right.style

            # Alignment
            if cell.alignment:
                format_obj.horizontal_align = cell.alignment.horizontal
                format_obj.vertical_align = cell.alignment.vertical
                format_obj.text_wrap = cell.alignment.wrap_text

            # Number format
            format_obj.number_format = cell.number_format

            return format_obj

        except Exception as e:
            logger.warning(f"Failed to extract cell format: {e}")
            return CellFormat()

    def _excel_color_to_hex(self, color: Any) -> Optional[str]:
        """Convert Excel color object to hex string."""
        try:
            if hasattr(color, "rgb") and color.rgb:
                rgb = color.rgb
                if len(rgb) == 8:  # ARGB format
                    return f"#{rgb[2:]}"  # Remove alpha channel
                elif len(rgb) == 6:  # RGB format
                    return f"#{rgb}"
            return None
        except Exception:
            return None

    def extract_google_sheets_formatting(
        self, sheets_service: Any, spreadsheet_id: str, sheet_name: str
    ) -> Dict[str, CellFormat]:
        """Extract formatting from Google Sheets.

        Args:
            sheets_service: Google Sheets API service
            spreadsheet_id: ID of the spreadsheet
            sheet_name: Name of the sheet

        Returns:
            Dict mapping cell coordinates to CellFormat objects
        """
        try:
            # Get sheet metadata including formatting
            request = sheets_service.spreadsheets().get(
                spreadsheetId=spreadsheet_id, includeGridData=True, ranges=[sheet_name]
            )

            response = request.execute()

            if "sheets" not in response:
                return {}

            cell_formats = {}

            for sheet in response["sheets"]:
                if sheet["properties"]["title"] == sheet_name:
                    if "data" in sheet:
                        for grid_data in sheet["data"]:
                            if "rowData" in grid_data:
                                self._process_google_sheets_rows(
                                    grid_data["rowData"], cell_formats
                                )

            logger.info(
                f"Extracted formatting for {len(cell_formats)} cells from Google Sheets"
            )
            return cell_formats

        except Exception as e:
            logger.error(f"Failed to extract Google Sheets formatting: {e}")
            return {}

    def _process_google_sheets_rows(
        self, row_data: List[Dict], cell_formats: Dict[str, CellFormat]
    ) -> None:
        """Process Google Sheets row data to extract formatting."""
        for row_idx, row in enumerate(row_data):
            if "values" in row:
                for col_idx, cell in enumerate(row["values"]):
                    if "effectiveFormat" in cell:
                        cell_ref = self._rc_to_a1(row_idx + 1, col_idx + 1)
                        cell_formats[cell_ref] = self._extract_google_cell_format(
                            cell["effectiveFormat"]
                        )

    def _extract_google_cell_format(self, effective_format: Dict) -> CellFormat:
        """Extract formatting from Google Sheets cell format."""
        format_obj = CellFormat()

        try:
            # Text format (font)
            if "textFormat" in effective_format:
                text_format = effective_format["textFormat"]

                if "fontFamily" in text_format:
                    format_obj.font_name = text_format["fontFamily"]
                if "fontSize" in text_format:
                    format_obj.font_size = text_format["fontSize"]
                if "bold" in text_format:
                    format_obj.font_bold = text_format["bold"]
                if "italic" in text_format:
                    format_obj.font_italic = text_format["italic"]
                if "underline" in text_format:
                    format_obj.font_underline = text_format["underline"]

                # Font color
                if "foregroundColor" in text_format:
                    format_obj.font_color = self._google_color_to_hex(
                        text_format["foregroundColor"]
                    )

            # Background color
            if "backgroundColor" in effective_format:
                format_obj.background_color = self._google_color_to_hex(
                    effective_format["backgroundColor"]
                )

            # Borders
            if "borders" in effective_format:
                borders = effective_format["borders"]
                if "top" in borders:
                    format_obj.border_top = self._google_border_style(borders["top"])
                if "bottom" in borders:
                    format_obj.border_bottom = self._google_border_style(
                        borders["bottom"]
                    )
                if "left" in borders:
                    format_obj.border_left = self._google_border_style(borders["left"])
                if "right" in borders:
                    format_obj.border_right = self._google_border_style(
                        borders["right"]
                    )

            # Alignment
            if "horizontalAlignment" in effective_format:
                format_obj.horizontal_align = self._translate_horizontal_alignment(
                    effective_format["horizontalAlignment"]
                )
            if "verticalAlignment" in effective_format:
                format_obj.vertical_align = self._translate_vertical_alignment(
                    effective_format["verticalAlignment"]
                )
            if "wrapStrategy" in effective_format:
                format_obj.text_wrap = effective_format["wrapStrategy"] == "WRAP"

            # Number format
            if "numberFormat" in effective_format:
                if "pattern" in effective_format["numberFormat"]:
                    format_obj.number_format = effective_format["numberFormat"][
                        "pattern"
                    ]

        except Exception as e:
            logger.warning(f"Failed to extract Google Sheets cell format: {e}")

        return format_obj

    def _google_color_to_hex(self, color: Dict) -> Optional[str]:
        """Convert Google Sheets color to hex string."""
        try:
            r = int(color.get("red", 0) * 255)
            g = int(color.get("green", 0) * 255)
            b = int(color.get("blue", 0) * 255)
            return f"#{r:02x}{g:02x}{b:02x}"
        except Exception:
            return None

    def _google_border_style(self, border: Dict) -> Optional[str]:
        """Extract border style from Google Sheets border object."""
        return border.get("style", "").lower() if border else None

    def _translate_horizontal_alignment(self, google_alignment: str) -> str:
        """Translate Google Sheets horizontal alignment to Excel format."""
        alignment_map = {
            "LEFT": "left",
            "CENTER": "center",
            "RIGHT": "right",
            "JUSTIFY": "justify",
        }
        return alignment_map.get(google_alignment, "general")

    def _translate_vertical_alignment(self, google_alignment: str) -> str:
        """Translate Google Sheets vertical alignment to Excel format."""
        alignment_map = {
            "TOP": "top",
            "MIDDLE": "center",
            "BOTTOM": "bottom",
            "JUSTIFY": "justify",
        }
        return alignment_map.get(google_alignment, "bottom")

    def _rc_to_a1(self, row: int, col: int) -> str:
        """Convert row/column numbers to A1 notation."""
        result = ""
        while col > 0:
            col -= 1
            result = chr(col % 26 + ord("A")) + result
            col //= 26
        return result + str(row)


class FormatApplier:
    """Applies formatting to spreadsheets."""

    def apply_excel_formatting(
        self, workbook: Any, sheet_name: str, cell_formats: Dict[str, CellFormat]
    ) -> None:
        """Apply formatting to Excel worksheet.

        Args:
            workbook: openpyxl workbook object
            sheet_name: Name of the worksheet
            cell_formats: Dict mapping cell coordinates to CellFormat objects
        """
        if not OPENPYXL_AVAILABLE:
            logger.warning("openpyxl not available - cannot apply Excel formatting")
            return

        try:
            worksheet = workbook[sheet_name]

            for cell_ref, format_obj in cell_formats.items():
                cell = worksheet[cell_ref]
                self._apply_excel_cell_format(cell, format_obj)

            logger.info(
                f"Applied formatting to {len(cell_formats)} cells in {sheet_name}"
            )

        except Exception as e:
            logger.error(f"Failed to apply Excel formatting: {e}")

    def _apply_excel_cell_format(self, cell: Any, format_obj: CellFormat) -> None:
        """Apply formatting to a single Excel cell."""
        try:
            # Font formatting
            if any(
                [
                    format_obj.font_name,
                    format_obj.font_size,
                    format_obj.font_bold,
                    format_obj.font_italic,
                    format_obj.font_underline,
                    format_obj.font_color,
                ]
            ):
                font = Font(
                    name=format_obj.font_name or cell.font.name,
                    size=format_obj.font_size or cell.font.size,
                    bold=format_obj.font_bold
                    if format_obj.font_bold is not None
                    else cell.font.bold,
                    italic=format_obj.font_italic
                    if format_obj.font_italic is not None
                    else cell.font.italic,
                    underline="single" if format_obj.font_underline else None,
                    color=format_obj.font_color[1:]
                    if format_obj.font_color
                    else None,  # Remove #
                )
                cell.font = font

            # Background color
            if format_obj.background_color:
                fill = PatternFill(
                    start_color=format_obj.background_color[1:],  # Remove #
                    end_color=format_obj.background_color[1:],
                    fill_type="solid",
                )
                cell.fill = fill

            # Alignment
            if any(
                [
                    format_obj.horizontal_align,
                    format_obj.vertical_align,
                    format_obj.text_wrap,
                ]
            ):
                alignment = Alignment(
                    horizontal=format_obj.horizontal_align or cell.alignment.horizontal,
                    vertical=format_obj.vertical_align or cell.alignment.vertical,
                    wrap_text=format_obj.text_wrap
                    if format_obj.text_wrap is not None
                    else cell.alignment.wrap_text,
                )
                cell.alignment = alignment

            # Number format
            if format_obj.number_format:
                cell.number_format = format_obj.number_format

        except Exception as e:
            logger.warning(f"Failed to apply cell formatting: {e}")

    def apply_google_sheets_formatting(
        self,
        sheets_service: Any,
        spreadsheet_id: str,
        sheet_id: int,
        cell_formats: Dict[str, CellFormat],
    ) -> None:
        """Apply formatting to Google Sheets.

        Args:
            sheets_service: Google Sheets API service
            spreadsheet_id: ID of the spreadsheet
            sheet_id: Internal sheet ID (not name)
            cell_formats: Dict mapping cell coordinates to CellFormat objects
        """
        try:
            # Step 1: Ensure sheet has enough rows/columns for all formatting
            max_row = max_col = 0
            for cell_ref in cell_formats.keys():
                row, col = self._a1_to_rc(cell_ref)
                max_row = max(max_row, row)
                max_col = max(max_col, col)

            if max_row > 0 or max_col > 0:
                # Expand sheet if necessary (add some buffer)
                target_rows = max(max_row + 10, 1000)  # At least 1000 rows
                target_cols = max(max_col + 5, 26)  # At least 26 columns

                expand_request = {
                    "updateSheetProperties": {
                        "properties": {
                            "sheetId": sheet_id,
                            "gridProperties": {
                                "rowCount": target_rows,
                                "columnCount": target_cols,
                            },
                        },
                        "fields": "gridProperties.rowCount,gridProperties.columnCount",
                    }
                }

                # Execute grid expansion first
                expand_body = {"requests": [expand_request]}
                sheets_service.spreadsheets().batchUpdate(
                    spreadsheetId=spreadsheet_id, body=expand_body
                ).execute()

                logger.info(
                    f"Expanded Google Sheets to {target_rows} rows x {target_cols} columns"
                )

            # Step 2: Apply formatting
            requests = []

            for cell_ref, format_obj in cell_formats.items():
                row, col = self._a1_to_rc(cell_ref)

                # Create format request
                format_request = self._create_google_format_request(
                    sheet_id, row, col, format_obj
                )
                if format_request:
                    requests.append(format_request)

            if requests:
                # Apply all formatting in batch
                body = {"requests": requests}
                sheets_service.spreadsheets().batchUpdate(
                    spreadsheetId=spreadsheet_id, body=body
                ).execute()

                logger.info(
                    f"Applied formatting to {len(cell_formats)} cells in Google Sheets"
                )

        except Exception as e:
            logger.error(f"Failed to apply Google Sheets formatting: {e}")

    def _create_google_format_request(
        self, sheet_id: int, row: int, col: int, format_obj: CellFormat
    ) -> Optional[Dict]:
        """Create Google Sheets format request for a single cell."""
        try:
            cell_format = {}

            # Text format (font)
            text_format = {}
            if format_obj.font_name:
                text_format["fontFamily"] = format_obj.font_name
            if format_obj.font_size:
                text_format["fontSize"] = format_obj.font_size
            if format_obj.font_bold is not None:
                text_format["bold"] = format_obj.font_bold
            if format_obj.font_italic is not None:
                text_format["italic"] = format_obj.font_italic
            if format_obj.font_underline is not None:
                text_format["underline"] = format_obj.font_underline
            if format_obj.font_color:
                text_format["foregroundColor"] = self._hex_to_google_color(
                    format_obj.font_color
                )

            if text_format:
                cell_format["textFormat"] = text_format

            # Background color
            if format_obj.background_color:
                cell_format["backgroundColor"] = self._hex_to_google_color(
                    format_obj.background_color
                )

            # Alignment
            if format_obj.horizontal_align:
                cell_format["horizontalAlignment"] = format_obj.horizontal_align.upper()
            if format_obj.vertical_align:
                cell_format["verticalAlignment"] = format_obj.vertical_align.upper()
            if format_obj.text_wrap is not None:
                cell_format["wrapStrategy"] = (
                    "WRAP" if format_obj.text_wrap else "OVERFLOW_CELL"
                )

            # Number format - Google Sheets requires both type and pattern
            if format_obj.number_format:
                # Determine the format type based on the pattern
                pattern = format_obj.number_format
                format_type = "NUMBER"  # Default

                if any(
                    date_indicator in pattern.lower()
                    for date_indicator in ["mm", "dd", "yyyy", "yy", "m/", "d/"]
                ):
                    format_type = "DATE"
                elif any(
                    time_indicator in pattern.lower()
                    for time_indicator in ["hh", "mm", "ss", "am", "pm"]
                ):
                    format_type = "TIME"
                elif "%" in pattern:
                    format_type = "PERCENT"
                elif "$" in pattern or "¤" in pattern:
                    format_type = "CURRENCY"
                elif "e" in pattern.lower():
                    format_type = "SCIENTIFIC"

                cell_format["numberFormat"] = {"type": format_type, "pattern": pattern}

            if not cell_format:
                return None

            return {
                "repeatCell": {
                    "range": {
                        "sheetId": sheet_id,
                        "startRowIndex": row - 1,
                        "endRowIndex": row,
                        "startColumnIndex": col - 1,
                        "endColumnIndex": col,
                    },
                    "cell": {"userEnteredFormat": cell_format},
                    "fields": "userEnteredFormat",
                }
            }

        except Exception as e:
            logger.warning(f"Failed to create format request: {e}")
            return None

    def _hex_to_google_color(self, hex_color: str) -> Dict:
        """Convert hex color to Google Sheets color format."""
        try:
            hex_color = hex_color.lstrip("#")
            r = int(hex_color[0:2], 16) / 255.0
            g = int(hex_color[2:4], 16) / 255.0
            b = int(hex_color[4:6], 16) / 255.0
            return {"red": r, "green": g, "blue": b}
        except Exception:
            return {"red": 0, "green": 0, "blue": 0}

    def _a1_to_rc(self, a1_ref: str) -> Tuple[int, int]:
        """Convert A1 notation to row/column numbers."""
        import re

        match = re.match(r"([A-Z]+)(\d+)", a1_ref)
        if not match:
            return 1, 1

        col_letters = match.group(1)
        row = int(match.group(2))

        col = 0
        for char in col_letters:
            col = col * 26 + (ord(char) - ord("A") + 1)

        return row, col


def preserve_formatting_during_conversion(
    source_formats: Dict[str, CellFormat],
    source_is_google: bool,
    target_is_google: bool,
) -> Dict[str, CellFormat]:
    """Translate formatting between Excel and Google Sheets formats.

    Args:
        source_formats: Formatting extracted from source
        source_is_google: Whether source is Google Sheets
        target_is_google: Whether target is Google Sheets

    Returns:
        Translated formatting for target platform
    """
    if source_is_google == target_is_google:
        # Same platform, no translation needed
        return source_formats

    translated_formats = {}

    for cell_ref, format_obj in source_formats.items():
        # Create a copy for translation
        translated_format = CellFormat(
            font_name=format_obj.font_name,
            font_size=format_obj.font_size,
            font_bold=format_obj.font_bold,
            font_italic=format_obj.font_italic,
            font_underline=format_obj.font_underline,
            font_color=format_obj.font_color,
            background_color=format_obj.background_color,
            border_top=format_obj.border_top,
            border_bottom=format_obj.border_bottom,
            border_left=format_obj.border_left,
            border_right=format_obj.border_right,
            border_color=format_obj.border_color,
            horizontal_align=format_obj.horizontal_align,
            vertical_align=format_obj.vertical_align,
            text_wrap=format_obj.text_wrap,
            number_format=format_obj.number_format,
        )

        # Perform platform-specific translations
        if source_is_google and not target_is_google:
            # Google Sheets → Excel translations
            translated_format = _translate_google_to_excel_format(translated_format)
        elif not source_is_google and target_is_google:
            # Excel → Google Sheets translations
            translated_format = _translate_excel_to_google_format(translated_format)

        translated_formats[cell_ref] = translated_format

    return translated_formats


def _translate_google_to_excel_format(format_obj: CellFormat) -> CellFormat:
    """Translate Google Sheets formatting to Excel equivalents."""
    # Font name translations
    if format_obj.font_name:
        font_mapping = {
            "Arial": "Arial",
            "Roboto": "Calibri",  # Google's default → Excel's default
            "Times New Roman": "Times New Roman",
        }
        format_obj.font_name = font_mapping.get(
            format_obj.font_name, format_obj.font_name
        )

    # Alignment translations
    if format_obj.horizontal_align:
        alignment_mapping = {"left": "left", "center": "center", "right": "right"}
        format_obj.horizontal_align = alignment_mapping.get(
            format_obj.horizontal_align, format_obj.horizontal_align
        )

    return format_obj


def _translate_excel_to_google_format(format_obj: CellFormat) -> CellFormat:
    """Translate Excel formatting to Google Sheets equivalents."""
    # Font name translations
    if format_obj.font_name:
        font_mapping = {
            "Calibri": "Arial",  # Excel's default → Google's compatible
            "Arial": "Arial",
            "Times New Roman": "Times New Roman",
        }
        format_obj.font_name = font_mapping.get(
            format_obj.font_name, format_obj.font_name
        )

    # Alignment translations
    if format_obj.horizontal_align:
        alignment_mapping = {
            "general": "left",  # Excel's default
            "left": "left",
            "center": "center",
            "right": "right",
        }
        format_obj.horizontal_align = alignment_mapping.get(
            format_obj.horizontal_align, format_obj.horizontal_align
        )

    return format_obj
