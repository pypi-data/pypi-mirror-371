"""Migrated spreadsheet conversion tests using centralized mocking utilities.

This demonstrates how the centralized mocking approach dramatically simplifies
spreadsheet conversion tests while maintaining full test coverage.
"""

import pytest
import tempfile
import shutil
from pathlib import Path
from unittest.mock import Mock, patch

from urarovite.utils.generic_spreadsheet import (
    convert_google_sheets_to_excel,
    convert_excel_to_google_sheets,
    convert_spreadsheet_format,
    _copy_google_sheets_to_google_sheets,
    _copy_excel_to_excel,
)
from urarovite.core.exceptions import ValidationError
from .fixtures import BaseTestCase
from .mock_manager import mock_manager


class TestConvertGoogleSheetsToExcel(BaseTestCase):
    """Test Google Sheets to Excel conversion."""

    def setup_method(self):
        """Set up test fixtures."""
        super().setup_method()
        self.temp_dir = Path(tempfile.mkdtemp())

    def teardown_method(self):
        """Clean up test fixtures."""
        if self.temp_dir.exists():
            shutil.rmtree(self.temp_dir)

    def test_successful_conversion(self):
        """Test successful Google Sheets to Excel conversion."""
        with mock_manager.spreadsheet_scenario(
            source_sheets=["Sheet1", "Sheet2"],
            target_sheets=["Sheet1", "Sheet2"],
            sheet_data={
                "Sheet1": [["A1", "B1"], ["A2", "B2"]],
                "Sheet2": [["C1", "D1"], ["C2", "D2"]],
            },
        ):
            result = convert_google_sheets_to_excel(
                google_sheets_url="https://docs.google.com/spreadsheets/d/test123/edit",
                excel_file_path=self.temp_dir / "output.xlsx",
                auth_credentials={"auth_secret": self.encoded_creds},
            )

            # Verify result
            assert result["success"] is True
            assert len(result["converted_sheets"]) == 2
            assert "Sheet1" in result["converted_sheets"]
            assert "Sheet2" in result["converted_sheets"]
            assert result["output_path"].endswith("output.xlsx")
            assert result["error"] is None

    def test_specific_sheets_conversion(self):
        """Test conversion of specific sheets only."""
        with mock_manager.spreadsheet_scenario(
            source_sheets=["Sheet1", "Sheet2", "Sheet3"],
            target_sheets=["Sheet1", "Sheet3"],
            sheet_data={
                "Sheet1": [["A1", "B1"]],
                "Sheet2": [["C1", "D1"]],
                "Sheet3": [["E1", "F1"]],
            },
        ):
            result = convert_google_sheets_to_excel(
                google_sheets_url="https://docs.google.com/spreadsheets/d/test123/edit",
                excel_file_path=self.temp_dir / "output.xlsx",
                auth_credentials={"auth_secret": self.encoded_creds},
                sheet_names=["Sheet1", "Sheet3"],
            )

            assert result["success"] is True
            assert len(result["converted_sheets"]) == 2
            assert "Sheet1" in result["converted_sheets"]
            assert "Sheet3" in result["converted_sheets"]
            assert "Sheet2" not in result["converted_sheets"]

    def test_conversion_with_errors(self):
        """Test conversion with some errors."""
        with mock_manager.spreadsheet_scenario(
            source_sheets=["Sheet1", "Sheet2"],
            target_sheets=["Sheet1"]
        ) as mocks:
            # Make one sheet fail
            mocks["source"].get_sheet_data.side_effect = [
                Mock(values=[["A1", "B1"]], sheet_name="Sheet1"),
                Exception("Failed to get Sheet2 data"),
            ]

            result = convert_google_sheets_to_excel(
                google_sheets_url="https://docs.google.com/spreadsheets/d/test123/edit",
                excel_file_path=self.temp_dir / "output.xlsx",
                auth_credentials={"auth_secret": self.encoded_creds},
            )

            assert result["success"] is True  # Partial success
            assert len(result["converted_sheets"]) == 1
            assert "Sheet1" in result["converted_sheets"]
            # Note: Individual sheet errors are logged as warnings, not returned in result
            assert result["error"] is None  # Overall operation succeeded

    def test_file_extension_handling(self):
        """Test that .xlsx extension is added if missing."""
        with mock_manager.spreadsheet_scenario():
            result = convert_google_sheets_to_excel(
                google_sheets_url="https://docs.google.com/spreadsheets/d/test123/edit",
                excel_file_path=self.temp_dir / "output",  # No extension
                auth_credentials={"auth_secret": self.encoded_creds},
            )

            assert result["success"] is True
            assert result["output_path"].endswith(".xlsx")


class TestConvertExcelToGoogleSheets(BaseTestCase):
    """Test Excel to Google Sheets conversion."""

    def setup_method(self):
        """Set up test fixtures."""
        super().setup_method()
        self.temp_dir = Path(tempfile.mkdtemp())

    def teardown_method(self):
        """Clean up test fixtures."""
        if self.temp_dir.exists():
            shutil.rmtree(self.temp_dir)

    def test_successful_conversion(self):
        """Test successful Excel to Google Sheets conversion."""
        # Create a test Excel file
        test_file = self.temp_dir / "input.xlsx"
        try:
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws["A1"] = "A1"
            ws["B1"] = "B1"
            ws["A2"] = "A2"
            ws["B2"] = "B2"
            wb.save(str(test_file))
        except ImportError:
            # Create a dummy file if openpyxl not available
            test_file.write_text("dummy excel content")

        with mock_manager.spreadsheet_scenario(
            source_sheets=["Sheet1"],
            sheet_data={"Sheet1": [["A1", "B1"], ["A2", "B2"]]},
        ):
            result = convert_excel_to_google_sheets(
                excel_file_path=self.temp_dir / "input.xlsx",
                google_sheets_url="https://docs.google.com/spreadsheets/d/target123/edit",
                auth_credentials={"auth_secret": self.encoded_creds},
            )

            assert result["success"] is True
            assert len(result["converted_sheets"]) == 1
            assert "Sheet1" in result["converted_sheets"]
            assert result["error"] is None

    def test_create_new_sheets_option(self):
        """Test creating new sheets in target."""
        # Create a test Excel file
        test_file = self.temp_dir / "input.xlsx"
        try:
            from openpyxl import Workbook

            wb = Workbook()
            ws1 = wb.active
            ws1.title = "NewSheet1"
            ws1["A1"] = "Data1"
            ws2 = wb.create_sheet("NewSheet2")
            ws2["A1"] = "Data2"
            wb.save(str(test_file))
        except ImportError:
            # Create a dummy file if openpyxl not available
            test_file.write_text("dummy excel content")

        with mock_manager.spreadsheet_scenario(
            source_sheets=["NewSheet1", "NewSheet2"],
            target_sheets=["Sheet"],  # Target has different sheets
        ) as mocks:
            result = convert_excel_to_google_sheets(
                excel_file_path=self.temp_dir / "input.xlsx",
                google_sheets_url="https://docs.google.com/spreadsheets/d/target123/edit",
                auth_credentials={"auth_secret": self.encoded_creds},
                create_new_sheets=True,
            )

            assert result["success"] is True
            # Should have created new sheets
            assert mocks["target"].create_sheet.call_count == 2


class TestConvertSpreadsheetFormat(BaseTestCase):
    """Test the main convert_spreadsheet_format function."""

    def setup_method(self):
        """Set up test fixtures."""
        super().setup_method()
        self.temp_dir = Path(tempfile.mkdtemp())

    def teardown_method(self):
        """Clean up test fixtures."""
        if self.temp_dir.exists():
            shutil.rmtree(self.temp_dir)

    def test_google_sheets_to_excel_detection(self):
        """Test auto-detection of Google Sheets to Excel conversion."""
        # Mock the underlying conversion functions to return success
        with patch(
            "urarovite.utils.enhanced_conversion.convert_with_full_formatting"
        ) as mock_enhanced_convert:
            mock_enhanced_convert.return_value = {
                "success": True,
                "conversion_type": "google_sheets_to_excel",
                "converted_sheets": ["Sheet1"],
                "formula_stats": {"total_formulas": 0},
                "formatting_stats": {"cells_with_formatting": 0},
                "error": None,
            }

            result = convert_spreadsheet_format(
                source="https://docs.google.com/spreadsheets/d/source123/edit",
                target=str(self.temp_dir / "output.xlsx"),
                auth_credentials={"auth_secret": self.encoded_creds},
            )

            assert result["success"] is True
            assert result["conversion_type"] == "google_sheets_to_excel"

    def test_excel_to_google_sheets_detection(self):
        """Test auto-detection of Excel to Google Sheets conversion."""
        # Create a dummy Excel file for the test
        test_file = self.temp_dir / "input.xlsx"
        test_file.touch()

        with patch(
            "urarovite.utils.enhanced_conversion.convert_with_full_formatting"
        ) as mock_enhanced_convert:
            mock_enhanced_convert.return_value = {
                "success": True,
                "conversion_type": "excel_to_google_sheets",
                "converted_sheets": ["Sheet1"],
                "formula_stats": {"total_formulas": 0},
                "formatting_stats": {"cells_with_formatting": 0},
                "error": None,
            }

            result = convert_spreadsheet_format(
                source=str(test_file),
                target="https://docs.google.com/spreadsheets/d/target123/edit",
                auth_credentials={"auth_secret": self.encoded_creds},
            )

            assert result["success"] is True
            assert result["conversion_type"] == "excel_to_google_sheets"

    def test_google_sheets_to_google_sheets_detection(self):
        """Test auto-detection of Google Sheets to Google Sheets copy."""
        with patch(
            "urarovite.utils.enhanced_conversion.convert_with_full_formatting"
        ) as mock_enhanced_convert:
            mock_enhanced_convert.return_value = {
                "success": True,
                "conversion_type": "google_sheets_to_google_sheets",
                "converted_sheets": ["Sheet1"],
                "formula_stats": {"total_formulas": 0},
                "formatting_stats": {"cells_with_formatting": 0},
                "error": None,
            }

            result = convert_spreadsheet_format(
                source="https://docs.google.com/spreadsheets/d/source123/edit",
                target="https://docs.google.com/spreadsheets/d/target123/edit",
                auth_credentials={"auth_secret": self.encoded_creds},
            )

            assert result["success"] is True
            assert result["conversion_type"] == "google_sheets_to_google_sheets"

    def test_excel_to_excel_detection(self):
        """Test auto-detection of Excel to Excel copy."""
        # Create dummy Excel files for the test
        input_file = self.temp_dir / "input.xlsx"
        output_file = self.temp_dir / "output.xlsx"
        input_file.touch()

        with patch(
            "urarovite.utils.enhanced_conversion.convert_with_full_formatting"
        ) as mock_enhanced_convert:
            mock_enhanced_convert.return_value = {
                "success": True,
                "conversion_type": "excel_to_excel",
                "converted_sheets": ["Sheet1"],
                "formula_stats": {"total_formulas": 0},
                "formatting_stats": {"cells_with_formatting": 0},
                "error": None,
            }

            result = convert_spreadsheet_format(
                source=str(input_file),
                target=str(output_file),
                auth_credentials={"auth_secret": self.encoded_creds},
            )

            assert result["success"] is True
            assert result["conversion_type"] == "excel_to_excel"


class TestCopyFunctions(BaseTestCase):
    """Test the copy functions for same-format operations."""

    def setup_method(self):
        """Set up test fixtures."""
        super().setup_method()
        self.temp_dir = Path(tempfile.mkdtemp())

    def teardown_method(self):
        """Clean up test fixtures."""
        if self.temp_dir.exists():
            shutil.rmtree(self.temp_dir)

    def test_copy_google_sheets_to_google_sheets(self):
        """Test copying between Google Sheets."""
        with mock_manager.spreadsheet_scenario():
            result = _copy_google_sheets_to_google_sheets(
                source_url="https://docs.google.com/spreadsheets/d/source123/edit",
                target_url="https://docs.google.com/spreadsheets/d/target123/edit",
                auth_credentials={"auth_secret": self.encoded_creds},
            )

            assert result["success"] is True
            assert "converted_sheets" in result

    def test_copy_excel_to_excel(self):
        """Test copying between Excel files."""
        with mock_manager.spreadsheet_scenario():
            result = _copy_excel_to_excel(
                source_path=str(self.temp_dir / "source.xlsx"),
                target_path=str(self.temp_dir / "target.xlsx"),
            )

            assert result["success"] is True
            assert "converted_sheets" in result

    def test_copy_excel_extension_handling(self):
        """Test Excel copy with extension handling."""
        with mock_manager.spreadsheet_scenario():
            result = _copy_excel_to_excel(
                source_path=str(self.temp_dir / "source.xlsx"),
                target_path=str(self.temp_dir / "target"),  # No extension
            )

            assert result["success"] is True
            assert result["output_path"].endswith(".xlsx")


class TestErrorHandling(BaseTestCase):
    """Test error handling in spreadsheet conversion."""

    def setup_method(self):
        """Set up test fixtures."""
        super().setup_method()
        self.temp_dir = Path(tempfile.mkdtemp())

    def teardown_method(self):
        """Clean up test fixtures."""
        if self.temp_dir.exists():
            shutil.rmtree(self.temp_dir)

    def test_convert_excel_to_google_sheets_error(self):
        """Test handling of Excel to Google Sheets conversion errors."""
        with mock_manager.spreadsheet_scenario() as mocks:
            # Make the source spreadsheet fail
            mocks["source"].__enter__.side_effect = Exception(
                "Excel file not found: /tmp/test.xlsx"
            )

            result = convert_excel_to_google_sheets(
                excel_file_path="/tmp/test.xlsx",
                google_sheets_url="https://docs.google.com/spreadsheets/d/target123/edit",
                auth_credentials={"auth_secret": self.encoded_creds},
            )

            assert result["success"] is False
            assert "Excel file not found: /tmp/test.xlsx" in result["error"]

    def test_convert_spreadsheet_format_unknown_type(self):
        """Test handling of missing files in conversion."""
        # This test checks error handling when files don't exist
        result = convert_spreadsheet_format(
            source=str(self.temp_dir / "input.xlsx"),
            target=str(self.temp_dir / "output.xlsx"),
            auth_credentials={"auth_secret": self.encoded_creds},
        )

        # Should fail due to missing file
        assert result["success"] is False
        assert "Excel file not found" in result["error"]


# Comparison of old vs new approach for spreadsheet conversion tests:

"""
OLD APPROACH (original test_spreadsheet_conversion.py):
=======================================================

@patch('urarovite.utils.generic_spreadsheet.SpreadsheetFactory.create_spreadsheet')
def test_successful_conversion(self, mock_factory):
    # Setup mocks using helper function
    mock_source = create_mock_spreadsheet()
    mock_source.get_metadata.return_value = Mock(sheet_names=["Sheet1", "Sheet2"])
    mock_source.get_sheet_data.side_effect = [
        Mock(values=[["A1", "B1"], ["A2", "B2"]], sheet_name="Sheet1"),
        Mock(values=[["C1", "D1"], ["C2", "D2"]], sheet_name="Sheet2")
    ]
    
    mock_target = create_mock_spreadsheet()
    mock_target.get_metadata.return_value = Mock(sheet_names=["Sheet"])
    
    mock_factory.side_effect = [mock_source, mock_target]

    # Execute conversion
    result = convert_google_sheets_to_excel(
        google_sheets_url="https://docs.google.com/spreadsheets/d/test123/edit",
        excel_file_path=self.temp_dir / "output.xlsx",
        auth_credentials={"auth_secret": "test_creds"}
    )

    # Verify result
    assert result["success"] is True
    assert len(result["converted_sheets"]) == 2
    assert "Sheet1" in result["converted_sheets"]
    assert "Sheet2" in result["converted_sheets"]
    assert result["output_path"].endswith("output.xlsx")
    assert result["error"] is None

LINES OF CODE: ~25 lines of setup + test logic


NEW APPROACH (this migrated file):
==================================

def test_successful_conversion(self):
    with mock_manager.spreadsheet_scenario(
        source_sheets=["Sheet1", "Sheet2"],
        sheet_data={
            "Sheet1": [["A1", "B1"], ["A2", "B2"]],
            "Sheet2": [["C1", "D1"], ["C2", "D2"]]
        }
    ) as mocks:
        result = convert_google_sheets_to_excel(
            google_sheets_url="https://docs.google.com/spreadsheets/d/test123/edit",
            excel_file_path=self.temp_dir / "output.xlsx",
            auth_credentials={"auth_secret": self.encoded_creds}
        )

        # Verify result
        assert result["success"] is True
        assert len(result["converted_sheets"]) == 2
        assert "Sheet1" in result["converted_sheets"]
        assert "Sheet2" in result["converted_sheets"]
        assert result["output_path"].endswith("output.xlsx")
        assert result["error"] is None

LINES OF CODE: ~18 lines total

REDUCTION: ~28% fewer lines, much cleaner setup, same functionality
"""
