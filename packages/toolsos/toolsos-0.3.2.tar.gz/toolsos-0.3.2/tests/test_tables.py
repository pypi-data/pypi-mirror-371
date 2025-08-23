import openpyxl
import pandas as pd
import pytest

from toolsos.huisstijl.tables.tables import write_table, write_table_from_dict

DATA = [["Noord", 1, 0.25], ["Zuid", 3, 0.75]]


COLUMNS = ["stadsdeel", "bewoners", "perc_bewoners"]


@pytest.fixture(scope="class")
def workbook(data=DATA, columns=COLUMNS):
    df1 = pd.DataFrame(data=DATA, columns=COLUMNS)

    filename = "temp/singe_table.xlsx"
    write_table(
        df1, file=filename, header_row=0, perc_pattern="perc_bewoners", style="new"
    )
    return openpyxl.load_workbook(filename)


class TestTable:
    @pytest.fixture(autouse=True)
    def setup(self, workbook):
        self.workbook = workbook
        self.ws = workbook.active

    def test_data(self, data=DATA):
        rows = [[c.value for c in row] for row in self.ws.iter_rows(2)]

        assert all(src == written for src, written in zip(DATA, rows))

    def test_data_header(self):
        header = [[c.value for c in row] for row in self.ws.iter_rows(2)]

        assert all(src == written for src, written in zip(DATA, header))

    def test_background_color_header(self):
        header_color = [
            [c.fill.fgColor.value for c in row] for row in self.ws.iter_rows(max_row=1)
        ]

        assert all([[c == "004699" for c in row] for row in header_color])

    # def test_text_color_header(self):
    #     header_text_color = [[c.font.color.value for in row] in self.ws]

    def test_formatting(self):
        perc_values = [
            [c.number_format for c in row]
            for row in self.ws.iter_rows(min_row=2, min_col=3)
        ]

        all([[c == "'0.0%'" for c in row] for row in perc_values])
