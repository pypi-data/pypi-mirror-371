#!/usr/bin/env python3
"""Baidu api backend
"""

import functools
import hashlib
import pathlib
import requests
import urllib
import urllib.parse
import openpyxl


class Baidu:
    DEFAULT_ENDPOINT = 'https://api.map.baidu.com'
    DEFAULT_API = '/weather/v1/'

    def __init__(self,
                 access_key=None,
                 secure_key=None,
                 endpoint=DEFAULT_ENDPOINT,
                 api=DEFAULT_API,
                 **kwargs):
        """Initialize Baidu with an API key and endpoint."""
        self.access_key = access_key
        self.secure_key = secure_key
        self.endpoint = endpoint
        self.api = api
        self.opts = kwargs

    def querystring(self, params):
        """get query string
        """
        return '&'.join(map(lambda x: f'{x[0]}={x[1]}', params.items()))

    def sign(self, params):
        """get signature for the request
        """
        if not params:
            query = {'ak': self.access_key}
        else:
            query = {**params, 'ak': self.access_key}

        path_query_str = f'{self.api}?{self.querystring(query)}'

        encode_str = urllib.parse.quote(path_query_str,
                                        safe='/:=&?#+!$,;\'@()*[]')
        signature = hashlib.md5(
            urllib.parse.quote_plus(
                (encode_str + self.secure_key)).encode()).hexdigest()
        return signature

    def download_weather_phenomenon_data(self, filepath):
        """download weather_phenomenon data
        """
        filepath = pathlib.Path(filepath).expanduser()
        url = self.opts.get(
            'phenomenon_reference_download_url',
            ('https://mapopen-website-wiki.cdn.bcebos.com/cityList/'
             '%E7%99%BE%E5%BA%A6%E5%9C%B0%E5%9B%BE%E5%A4%A9%E6'
             '%B0%94%E5%8F%96%E5%80%BC%E5%AF%B9%E7%85%A7%E8%A1'
             '%A8(0410).xlsx'))

        filepath.parent.mkdir(parents=True, exist_ok=True)

        with requests.get(url, timeout=5, stream=True) as resp:
            resp.raise_for_status()
            with open(filepath, 'wb') as fout:
                for chunk in resp.iter_content(chunk_size=8192):
                    fout.write(chunk)

    @functools.cache
    def weather_phenomenon_data(self, filepath=None):
        """load weather_phenomenon.txt data
        """
        if not filepath:
            filepath = pathlib.Path(
                self.opts.get(
                    'phenomenon_reference_path',
                    '~/.config/yycli/data/weather_phenomenon_reference.xlsx'))
        else:
            filepath = pathlib.Path(filepath)

        if not filepath.expanduser().is_file():
            self.download_weather_phenomenon_data(filepath.as_posix())

        workbook = openpyxl.load_workbook(filepath.expanduser())
        worksheet = workbook.get_sheet_by_name('天气现象')
        rows = worksheet.iter_rows(2, worksheet.max_row, 1,
                                   worksheet.max_column)
        sheet = list(map(lambda x: [cell.value for cell in x], rows))
        return sheet

    def download_district_id_data(self, savepath):
        """download district_id data
        """
        targetpath = pathlib.Path(savepath).expanduser()
        url = self.opts.get('district_id_reference_download_url',
                            ('https://mapopen-website-wiki.bj.bcebos.com'
                             '/cityList/weather_district_id.csv'))

        targetpath.parent.mkdir(parents=True, exist_ok=True)

        with requests.get(url, timeout=5, stream=True) as resp:
            resp.raise_for_status()
            with open(targetpath, 'wb') as fout:
                for chunk in resp.iter_content(chunk_size=8192):
                    fout.write(chunk)

    @functools.cache
    def weather_district_data(self, filepath=None):
        """load weather_district_id.csv data
        """
        if not filepath:
            filepath = pathlib.Path(
                self.opts.get('district_id_reference_path',
                              '~/.config/yycli/data/weather_district_id.csv'))

        targetpath = pathlib.Path(filepath).expanduser()

        if not targetpath.is_file():
            self.download_district_id_data(filepath.as_posix())

        with open(targetpath, 'r', encoding='utf-8') as fin:
            sheet = list(map(lambda x: x.strip().split(','), fin.readlines()))
            weather_data = sheet[1:]
            return weather_data

    @functools.cache
    def lookup_weather_district_by_id(self, district_id):
        """search district in weather_district_data by district id
        """
        for row in self.weather_district_data():
            if row[5] == district_id:
                return row
        return None

    @functools.cache
    def lookup_weather_district_by_text(self, district):
        """search district in weather_district_data by district text
        """
        for row in self.weather_district_data():
            if row[4].startswith(district):
                return row
        return None

    def resolve_query(self, query: str):
        """resolve district id
        """
        if query.isnumeric():
            # check if district_id valid
            if not self.lookup_weather_district_by_id(query):
                # invalid district id, ignore
                return None
            return query
        row = self.lookup_weather_district_by_text(query)
        if row:
            return row[5]
        return None

    def get_weather_data(self, district_id, data_type='all'):
        """Get weather data for a district by its ID.

        Args:
            district_id (str): District ID.
            data_type (str): Type of data to return ('all' or 'forecast').

        Returns:
            dict: Weather data in JSON format.
        """
        req_params = {
            'district_id': district_id,
            'data_type': data_type,
        }
        signature = self.sign(req_params)
        params = {**req_params, 'ak': self.access_key, 'sn': signature}
        response = requests.get((f'{self.endpoint}{self.api}'
                                 f'?{self.querystring(params)}'),
                                timeout=5)
        response.raise_for_status()
        return response.json()

    def query_weather(self, query: str):
        """Query weather data based on a district ID or name.

        Args:
            query (str): District ID or name.

        Returns:
            dict: Weather data in JSON format.
        """
        district_id = self.resolve_query(query)
        if not district_id:
            raise ValueError(f"Cannot resolve query: {query}")
        data = self.get_weather_data(district_id)
        return {
            'area': data['result']['location']['name'],
            'phenomenon': data['result']['now']['text'],
            'temperature': data['result']['now']['temp'],
            'wind_direction': data['result']['now']['wind_dir'],
            'wind_class': data['result']['now']['wind_class'],
        }
