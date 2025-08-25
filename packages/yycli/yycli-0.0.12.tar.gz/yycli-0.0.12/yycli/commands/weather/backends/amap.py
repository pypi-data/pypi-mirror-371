#!/usr/bin/env python3
"""Amap api backend
"""
import requests
import pathlib
import functools
import zipfile
import openpyxl


class Amap:
    """Amap api backend
    """
    DEFAULT_ENDPOINT = 'https://restapi.amap.com'
    DEFAULT_API = '/v3/weather/weatherInfo'

    def __init__(self,
                 key=None,
                 endpoint=DEFAULT_ENDPOINT,
                 api=DEFAULT_API,
                 **kwargs):
        """Initialize Amap with an API key."""
        if key is None:
            raise ValueError("API key must be provided")
        self.key = key
        self.endpoint = endpoint
        self.api = api
        self.opts = kwargs

    def download_adcode_data(self, savepath):
        """
        """
        url = self.opts.get(
            'adcode_reference_download_url',
            ('https://a.amap.com'
             '/lbs/static/code_resource/AMap_adcode_citycode.zip'))
        targetpath = pathlib.Path(savepath).expanduser()
        targetpath.parent.mkdir(parents=True, exist_ok=True)
        with requests.get(url, stream=True) as response:
            response.raise_for_status()
            with open(targetpath, 'wb') as file:
                for chunk in response.iter_content(chunk_size=8192):
                    file.write(chunk)

    @functools.cache
    def get_adcode_data(self, filepath=None):
        """Get adcode data.
        """
        if not filepath:
            filepath = pathlib.Path(
                self.opts.get('adcode_reference_path',
                              '~/.config/yycli/data/AMap_adcode_citycode.zip'))
        targetpath = pathlib.Path(filepath).expanduser()
        if not targetpath.is_file():
            self.download_adcode_data(targetpath)

        with zipfile.ZipFile(targetpath, 'r') as zip_ref:
            zip_ref.extract('AMap_adcode_citycode.xlsx', targetpath.parent)
        adcode_file = targetpath.parent / 'AMap_adcode_citycode.xlsx'
        if not adcode_file.is_file():
            raise FileNotFoundError(f"Adcode file not found: {adcode_file}")
        workbook = openpyxl.load_workbook(adcode_file)
        worksheet = workbook.active
        if not worksheet:
            raise ValueError("Worksheet is empty or not found in the workbook")
        rows = worksheet.iter_rows(2, worksheet.max_row, 1,
                                   worksheet.max_column)
        sheet = list(map(lambda x: [cell.value for cell in x], rows))
        return sheet

    @functools.cache
    def lookup_adcode_by_adcode(self, adcode: str):
        """Lookup adcode by adcode.

        Args:
            adcode (str): The adcode to lookup.

        Returns:
            matched adcode data or None if not found.
        """
        adcode_data = self.get_adcode_data()
        for row in adcode_data:
            if row[1] == adcode:
                return row[1]
        return None

    @functools.cache
    def lookup_adcode_by_citycode(self, citycode: str):
        """Lookup adcode by citycode.

        Args:
            citycode (str): The citycode to lookup.

        Returns:
            matched adcode data or None if not found.
        """
        adcode_data = self.get_adcode_data()
        for row in adcode_data:
            if row[2] == citycode:
                return row[1]
        return None

    @functools.cache
    def lookup_adcode_by_name(self, name: str):
        """Lookup adcode by city name.

        Args:
            name (str): The city name to lookup.

        Returns:
            matched adcode data or None if not found.
        """
        adcode_data = self.get_adcode_data()
        for row in adcode_data:
            if row[0].startswith(name):
                return row[1]
        return None

    @functools.cache
    def resolve_query(self, query: str):
        """Resolve a query name to an adcode.
        """
        if query.isnumeric():
            adcode = self.lookup_adcode_by_adcode(query)
            if adcode:
                return adcode
            adcode = self.lookup_adcode_by_citycode(query)
            if adcode:
                return adcode
            return None
        return self.lookup_adcode_by_name(query)

    def get_weather_data(self, city=440404, extensions='base', output='json'):
        """Get weather data for a city.

        Args:
            city (str): City code or name.
            extensions (str): Type of weather data to return ('base' or 'all').
            output (str): Format of the output ('json' or 'xml').

        Returns:
            dict: Weather data in JSON format.
        """
        params = {
            'key': self.key,
            'city': str(city),
            'extensions': extensions,
            'output': output
        }
        response = requests.get(f'{self.endpoint}{self.api}', params=params)
        response.raise_for_status()
        return response.json()

    def query_weather(self, query):
        """Query weather data based on a query string.
        """
        adcode = self.resolve_query(query)
        if not adcode:
            raise ValueError(f"Cannot resolve query: {query}")

        data = self.get_weather_data(int(adcode),
                                     extensions='base',
                                     output='json')
        return {
            'area': f'{data["lives"][0]["city"]}',
            'phenomenon': f'{data["lives"][0]["weather"]}',
            'temperature': f'{data["lives"][0]["temperature"]}',
            'wind_direction': f'{data["lives"][0]["winddirection"]}风',
            'wind_class': f'{data["lives"][0]["windpower"]}级',
        }
