import json
from dataclasses import dataclass
from pathlib import Path
from typing import TYPE_CHECKING

import openpyxl

from pbi_core import LocalReport
from pbi_core.static_files import Layout

if TYPE_CHECKING:
    from _typeshed import StrPath


@dataclass
class StaticElement:
    xpath: list[str | int]
    field: str
    text: str


class StaticElements:
    static_elements: dict[str, list[StaticElement]]

    def __init__(self) -> None:
        self.static_elements = {}

    def to_csv(self) -> None:
        pass

    def to_excel(self, path: "StrPath") -> None:
        wb = openpyxl.Workbook()
        for object_type, objects in self.static_elements.items():
            ws = wb.create_sheet(object_type)
            for j, name in enumerate(["xpath", "field", "default"]):
                ws.cell(1, j + 1).value = name
            for i, obj in enumerate(objects):
                ws.cell(2 + i, 1).value = json.dumps(obj.xpath)
                ws.cell(2 + i, 2).value = obj.field
                ws.cell(2 + i, 3).value = obj.text
        wb.remove(wb["Sheet"])
        wb.save(path)


def get_static_elements(layout: Layout) -> StaticElements:
    ret = StaticElements()
    # for section in layout.find_all(Section):
    #     ret.static_elements.setdefault("section", []).append(
    #         StaticElement(
    #             xpath=section.get_xpath(),
    #             field="displayName",
    #             text=section.displayName,
    #         ),
    #     )
    return ret


def set_static_elements(translation_path: "StrPath", pbix_path: "StrPath") -> None:
    wb = openpyxl.load_workbook(translation_path)
    languages: list[str] = [str(x) for x in next(iter(wb.worksheets[0].values))[3:]]
    processing: dict[str, list[StaticElement]] = {}
    for ws in wb.worksheets:
        for row in list(ws.values)[1:]:
            for i, language in enumerate(languages):
                processing.setdefault(language, []).append(
                    StaticElement(json.loads(str(row[0])), str(row[1]), str(row[3 + i])),
                )
    for language, static_elements in processing.items():
        pbix = LocalReport.load_pbix(pbix_path)
        for static_element in static_elements:
            node = pbix.static_files.layout.find_xpath(static_element.xpath)
            setattr(node, static_element.field, static_element.text)
        out_path = f"{Path(pbix_path).with_suffix('').absolute().as_posix()}_{language}.pbix"
        pbix.save_pbix(out_path)
